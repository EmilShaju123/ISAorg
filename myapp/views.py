import base64
import urllib
from collections import Counter
from datetime import datetime, timedelta, date
from sqlite3 import IntegrityError

from dateutil.relativedelta import relativedelta
from django.db import transaction
from django.db.models.functions import TruncMonth
from django.urls import reverse
from django.utils import timezone
import cv2
from bs4 import BeautifulSoup
from django.contrib import messages
from django.db.models import Q, Sum, Min, Max, Case, When
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import get_template, render_to_string
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from xhtml2pdf import pisa
from .forms import *
from .models import *
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseNotFound, JsonResponse
import num2words
from django.contrib.auth import authenticate, login as auth_login, logout
from django.contrib.auth.forms import UserCreationForm

# Create your views here.

def register(request):
    if 'username' in request.session:
        if request.method == 'POST':
            form = UserCreationForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'User registered successfully!')
                return render(request, 'admin/user/register.html', {'form': form})
        else:
            form = UserCreationForm()
        return render(request, 'admin/user/register.html', {'form': form})
    else:
        return redirect('login')

def staffindex(request):
    return render(request,'staffindex.html')

def login(request):
    if request.method == 'POST':
        form = logform(request.POST)
        if form.is_valid():
            uname = form.cleaned_data['uname']
            pas = form.cleaned_data['pas']
            user = authenticate(request, username=uname, password=pas)
            if user is not None:
                auth_login(request, user)
                # Set session variables
                request.session['username'] = uname
                request.session['is_superuser'] = user.is_superuser
                if user.is_superuser:
                    return redirect('adminhome')
                else:
                    return redirect('staffhome')
            else:
                messages.error(request, 'Incorrect username or password')
        else:
            messages.error(request, 'Invalid form data')
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    request.session.flush()
    messages.success(request, "You have been logged out.")
    return HttpResponseRedirect('/')

def userview(request):
    if 'username' in request.session:
        users = User.objects.filter(is_superuser=False,is_active=True)
        return render(request,'admin/user/viewuser.html',{'users':users})
    else:
        return redirect('login')

def userdel(request, username):
    """Deactivates a user account."""
    if 'username' not in request.session:
        return redirect('login')

    try:
        user = User.objects.get(username=username)
        user.is_active = False
        user.save()
        messages.success(request, f"User '{username}' deactivated successfully.")
        return redirect('userview')
    except User.DoesNotExist:
        messages.error(request, f"User '{username}' not found.")
        return redirect('userview')

def adminhome(request):
    if request.user.is_authenticated:
        return render(request, 'admin/adminhome.html', {'username': request.user.username})
    else:
        messages.info(request, "Please log in to access the admin home page.")
        return redirect('login')

def staffhome(request):
    if request.user.is_authenticated:
        return render(request, 'staff/staffhome.html', {'username': request.user.username})
    else:
        messages.info(request, "Please log in to access the staff home page.")
        return redirect('login')

def truckreg(request):
    if 'username' not in request.session:
        return redirect('login')
    today = datetime.today().strftime('%Y-%m-%d')
    transporters = transportermodel.objects.filter(checks=True)
    available_drivers = drivermodel.objects.filter(checks=True)
    if request.method == 'POST':
        code = request.POST.get('code')
        if truckmodel.objects.filter(code=code).exists():
            messages.error(request, 'Truck with this code already exists!')
            return render(request, 'staff/register/truckreg.html', {"data": transporters, "data1": available_drivers, 'today': today})
        else:
            driverid = request.POST.get('driver')
            driver = drivermodel.objects.get(driver=driverid)
            transid = request.POST.get('trans')
            truckno = request.POST.get('truckno')
            trans = transportermodel.objects.get(trans=transid)
            feet = request.POST.get('feet')
            emi = request.POST.get('emi') or 0
            tenure = request.POST.get('tenure') or None
            truck=truckmodel.objects.create(code=code, trans=trans, feet=feet, driver=driver, truckno=truckno, emi=emi, tenure=tenure)
            codes=truckmodel.objects.get(code=truck.code)
            if DriverAssignment.objects.filter(driver=driver).exists():
                assignment = DriverAssignment.objects.get(driver=driver)
                assignment.truck = codes
                assignment.save()
            else:
                DriverAssignment.objects.create(driver=driver, truck=codes, assignment_date=today)

            request.session['truck_registered'] = True
            return redirect('truckreg')
    else:
        if 'truck_registered' in request.session:
            messages.success(request, 'Truck registered successfully!')
            del request.session['truck_registered']
        return render(request, 'staff/register/truckreg.html', {"data": transporters, "data1": available_drivers, 'today': today})

def driver(request):
    if 'username' not in request.session:
        return redirect('login')
    today = date.today().strftime('%Y-%m-%d')
    if request.method == 'POST':
        try:
            driver_name = request.POST.get('driver')
            salary = request.POST.get('salary') or 0
            doj = request.POST.get('doj')
            if not driver_name:
                messages.error(request, 'Driver name is required!')
                return render(request, 'staff/register/driver.html', {'today': today})
            if drivermodel.objects.filter(driver=driver_name.upper()).exists():
                messages.error(request, 'Driver already exists!')
                return render(request, 'staff/register/driver.html', {'today': today})
            driver_instance = drivermodel.objects.create(driver=driver_name.upper(), salary=salary)
            DriverAssignment.objects.create(driver=driver_instance, assignment_date=doj)
            request.session['driver_added'] = True
            messages.success(request, 'Driver added successfully!')
            return redirect('driver')
        except IntegrityError:
            messages.error(request, 'Driver already exists!')
            return render(request, 'staff/register/driver.html', {'today': today})
        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'staff/register/driver.html', {'today': today})
    else:
        if 'driver_added' in request.session:
            del request.session['driver_added']
        return render(request, 'staff/register/driver.html', {'today': today})

def shift(request):
    if 'username' not in request.session:
        return redirect('login')

    if request.method == 'POST':
        try:
            trip = request.POST.get('trip')
            if not trip:
                messages.error(request, 'Trip is required!')
                return render(request, 'staff/register/shift.html')

            if shiftmodel.objects.filter(trip=trip).exists():
                messages.error(request, 'Trip already exists!')
                return render(request, 'staff/register/shift.html')

            shiftmodel.objects.create(trip=trip)
            request.session['shift_added'] = True
            return redirect('shift')
        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'staff/register/shift.html')
    else:
        if 'shift_added' in request.session:
            messages.success(request, 'Trip added successfully!')
            del request.session['shift_added']
        return render(request, 'staff/register/shift.html')

def transporterreg(request):
    if 'username' not in request.session:
        return redirect('login')

    if request.method == 'POST':
        try:
            trans = request.POST.get('trans')
            if not trans:
                messages.error(request, 'Transporter name is required!')
                return render(request, 'staff/register/transporterreg.html')

            if transportermodel.objects.filter(trans=trans).exists():
                messages.error(request, 'Transporter already exists!')
                return render(request, 'staff/register/transporterreg.html')

            transportermodel.objects.create(trans=trans)
            messages.success(request, 'Transporter registered successfully!')
            return redirect('transporterreg')
        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'staff/register/transporterreg.html')
    else:
        return render(request, 'staff/register/transporterreg.html')

def party(request):
    if 'username' not in request.session:
        return redirect('login')

    if request.method == 'POST':
        try:
            party = request.POST.get('party').strip()
            short = request.POST.get('short').strip()
            add = request.POST.get('add').strip()

            if not party or not add or not short:
                messages.error(request, 'Please fill in all fields!')
                return render(request, 'staff/register/party.html')

            if partymodel.objects.filter(party=party).exists():
                messages.error(request, 'Party already exists!')
                return render(request, 'staff/register/party.html')

            partymodel.objects.create(party=party, add=add, short=short)
            request.session['party_added'] = True
            messages.success(request, 'Party added successfully!')
            return redirect('party')
        except Exception as e:
            messages.error(request, str(e))
    return render(request, 'staff/register/party.html')

def place(request):
    if 'username' not in request.session:
        return redirect('login')

    if request.method == 'POST':
        try:
            place = request.POST.get('place')

            if not place:
                messages.error(request, 'Place is required!')
                return render(request, 'staff/register/place.html')

            if placemodel.objects.filter(place=place).exists():
                messages.error(request, 'Place already exists!')
                return render(request, 'staff/register/place.html')

            placemodel.objects.create(place=place)
            request.session['place_added'] = True
            return redirect('place')
        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'staff/register/place.html')
    else:
        if 'place_added' in request.session:
            messages.success(request, 'Place added successfully!')
            del request.session['place_added']
        return render(request, 'staff/register/place.html')

def battareg(request):
    if 'username' not in request.session:
        return redirect('login')

    existing_places = list(battaregmodel.objects.values_list('place', flat=True))
    place_counts = Counter(existing_places)
    places_to_exclude = [place for place, count in place_counts.items() if count == 2]
    places = placemodel.objects.exclude(place__in=places_to_exclude).filter(checks=True)

    if request.method == 'POST':
        try:
            place = placemodel.objects.get(place=request.POST.get('place'))
            battaregmodel.objects.create(
                place=place,
                feet=request.POST.get('feet'),
                batta=request.POST.get('batta')
            )
            messages.success(request, 'Batta rate created successfully!')
            return redirect('battareg')
        except placemodel.DoesNotExist:
            messages.error(request, 'Place not found')
        except Exception as e:
            messages.error(request, str(e))

    return render(request, 'staff/register/battareg.html', {'place': places})

def triporder(request):
    if 'username' not in request.session:
        return redirect('login')

    x = placemodel.objects.filter(checks=True)
    y = shiftmodel.objects.filter(checks=True)
    z = partymodel.objects.filter(checks=True)
    a = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
    d=drivermodel.objects.exclude(driver='OTHER')
    today_date = date.today().strftime('%Y-%m-%d')

    if 'ser' in request.POST:
        codes = request.POST.get('codes')
        try:
            trip = tripmodel.objects.filter(code=codes).latest('id')
            latest_trip = tripmodel.objects.filter(code=codes).exclude(place='SAME').latest('id')
            return render(request, 'staff/ISA/trip/triporder.html',
                          {'trip': trip, 'd3': a, 'today': today_date, "d": x, "d1": y, "d2": z, 'driver': d,'place':latest_trip.place,
                           'code': codes})
        except tripmodel.DoesNotExist:
            truck = truckmodel.objects.get(code=codes)
            trips = {'code': codes, 'driver': truck.driver, 'feet': truck.feet}
            return render(request, 'staff/ISA/trip/triporder.html',
                          {'trips': trips, 'd3': a, 'today': today_date, "d": x, "d1": y, "d2": z, 'driver': d,'code': codes})

    if 'reg' in request.POST:
        try:
            codeid = request.POST.get('code')
            code = truckmodel.objects.get(code=codeid)
            dat = request.POST.get('dat')
            dis = request.POST.get('dis') or 0
            disqnt = request.POST.get('disqnt') or 0
            partyid = request.POST.get('party')
            party = partymodel.objects.get(party=partyid)
            placeid = request.POST.get('place')
            driver = drivermodel.objects.get(driver=request.POST.get('driver'))
            feet = request.POST.get('feet')
            place = placemodel.objects.get(place=placeid)
            tripid = request.POST.get('trip')
            trip = shiftmodel.objects.filter(checks=True).get(trip=tripid)
            bno = request.POST.get('bno') or 0
            halt = request.POST.get('halt') or 0
            hire = request.POST.get('hire') or 0
            cont = request.POST.get('cont') or 'NIL'
            com = request.POST.get('com') or 0
            created_by = request.user
            trip_order = tripmodel.objects.create(
                code=code,
                driver=driver,
                dat=dat,
                party=party,
                place=place,
                trip=trip,
                bno=bno,
                dis=dis,
                disqnt=disqnt,
                created_by=created_by,
                halt=halt,
                hire=hire,
                cont=cont,
                com=com,
                feet=feet
            )
            if trip_order.feet == '20R' or trip_order.feet == '40R' or trip_order.feet == '2X20R':
                trip_order.batta = int(trip_order.hire) * 0.18
                trip_order.save()
            else:
                try:
                    if ((trip_order.feet == '2X20' or trip_order.feet == '3X20' or trip_order.feet == '4X20' or trip_order.
                            feet == '5X20' or trip_order.feet == '6X20' or trip_order.feet == '7X20' or trip_order.
                            feet == '8X20' or trip_order.feet == '9X20' or trip_order.feet == '10X20' or trip_order.
                            feet == '11X20' or trip_order.feet == '12X20' or trip_order.feet == '2X40' or trip_order.
                            feet == '3X40' or trip_order.feet == '4X40' or trip_order.feet == '5X40' or trip_order.
                            feet == '6X40' or trip_order.feet == '7X40' or trip_order.feet == '8X40' or trip_order.
                            feet == '9X40' or trip_order.feet == '10X40' or trip_order.feet == '11X40' or trip_order.
                            feet == '12X40' or trip_order.feet == '13X40' or trip_order.feet == '14X40' or trip_order.
                            feet == '15X40' or trip_order.feet == '16X40')or trip_order.feet == '17X40' or trip_order.
                            feet == '18X40' or trip_order.feet == '19X40' or trip_order.feet == '20X40'):
                        bata = battaregmodel.objects.get(place=placeid, feet='40')
                    else:
                        bata = battaregmodel.objects.get(place=placeid, feet=trip_order.feet)
                    trip_order.batta = bata.batta
                    trip_order.save()
                except battaregmodel.DoesNotExist:
                    trip_order.batta = 0
                    trip_order.save()
                    messages.warning(request, "Batta rate not found for this place and feet")
                    return render(request, 'staff/ISA/trip/triporder.html', {"d": x, "d1": y, "d2": z, "d3": a})
                request.session['trip_order_created'] = True
            messages.success(request, 'Trip order created successfully!')
            return redirect('triporder')
        except truckmodel.DoesNotExist:
            messages.error(request, "Truck not found")
        except partymodel.DoesNotExist:
            messages.error(request, "Party not found")
        except placemodel.DoesNotExist:
            messages.error(request, "Place not found")
        except shiftmodel.DoesNotExist:
            messages.error(request, "Trip not found")
        except battaregmodel.DoesNotExist:
            messages.error(request, "Batta rate not found")
        except Exception as e:
            messages.error(request, str(e))
        return render(request, 'staff/ISA/trip/triporder.html', {"d": x, "d1": y, "d2": z, "d3": a})
    else:
        if 'trip_order_created' in request.session:
            del request.session['trip_order_created']
        return render(request, 'staff/ISA/trip/triporder.html',
                      {"d": x, "d1": y, "d2": z, "d3": a, 'today': today_date, 'driver': d})

def triporderupdate(request):
    if 'username' not in request.session:
        return redirect('login')
    isa_trucks = truckmodel.objects.filter(trans__in=['ISA', 'THOLADAN', 'ISA TRANS'])
    filter_params = request.session.get('filter_params')
    data_submitted = False
    if 'ser1' in request.POST or filter_params:
        try:
            if 'ser1' in request.POST:
                data_submitted = True
                frdate = request.POST.get('frdate')
                todate = request.POST.get('todate')
                code = request.POST.get('code')
                cont = request.POST.get('cont')
                request.session['filter_params'] = {
                    'frdate': frdate,
                    'todate': todate,
                    'code': code,
                    'cont': cont,
                }
            else:
                frdate = filter_params.get('frdate')
                todate = filter_params.get('todate')
                code = filter_params.get('code')
                cont = filter_params.get('cont')
            if (not frdate or frdate == '') and (not todate or todate == '') and (code == '' or code == '----------------------------------Select---------------------------------------') and (not cont or cont == ''):
                messages.error(request, 'Select any filters')
                return render(request, 'staff/ISA/trip/triporderupdate.html', {'code': isa_trucks, 'frdate': frdate, 'todate': todate, 'truck': code})
            filter_q = Q()
            if code and code != '----------------------------------Select---------------------------------------':
                filter_q &= Q(code=code)
            if cont:
                cont_list = [c.strip() for c in cont.split(',')]
                q_objects = Q()
                for c in cont_list:
                    q_objects |= Q(cont__icontains=c)
                filter_q &= q_objects
            if frdate and todate:
                filter_q &= Q(dat__range=[frdate, todate])
            trips = tripmodel.objects.filter(filter_q).order_by('dat')
            if trips.exists():
                combined_data = []
                diesel = 0
                hire = 0
                halt = 0
                batta = 0
                battatotal = 0
                expense = 0
                for trip in trips:
                    trans_data = truckmodel.objects.filter(code=trip.code, trans__in=['ISA', 'THOLADAN', 'ISA TRANS']).first()
                    if trans_data:
                        diesel += trip.dis
                        halt += trip.halt
                        hire += trip.hire
                        batta += trip.batta
                        battatotal += trip.battatotal
                        expense += trip.exptotal
                        combined_data.append({'trip': trip})
                return render(request, 'staff/ISA/trip/triporderupdate.html', {'ser1': combined_data, 'diesel': diesel, 'hire': hire, 'code': isa_trucks, 'halt': halt, 'batta': batta, 'battatotal': battatotal, 'frdate': frdate, 'todate': todate, 'truck': code, 'expense': expense,'data_submitted': data_submitted})
            else:
                messages.error(request, 'No trips found.')
                return render(request, 'staff/ISA/trip/triporderupdate.html', {'code': isa_trucks, 'frdate': frdate, 'todate': todate, 'truck': code})
        except tripmodel.DoesNotExist:
            messages.error(request, 'No trips found.')
            return render(request, 'staff/ISA/trip/triporderupdate.html', {'code': isa_trucks, 'frdate': frdate, 'todate': todate, 'truck': code})
        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'staff/ISA/trip/triporderupdate.html', {'code': isa_trucks, 'frdate': frdate, 'todate': todate})
    return render(request, 'staff/ISA/trip/triporderupdate.html', {'code': isa_trucks})

def isaup(request, id):
    if 'username' not in request.session:
        return redirect('login')

    party = partymodel.objects.all()
    place = placemodel.objects.all()
    driver=drivermodel.objects.exclude(driver='OTHER')
    type = shiftmodel.objects.filter(checks=True)

    try:
        expense = tripmodel.objects.get(pk=id)
    except tripmodel.DoesNotExist:
        messages.error(request, "Trip not found.")
    bill = billmodel.objects.filter(bno=expense.bno)
    batta = battamodel.objects.filter(id=expense.battaid)

    if request.method == 'POST':
        try:
            expense.party = partymodel.objects.get(pk=request.POST.get('party'))
            expense.place = placemodel.objects.get(pk=request.POST.get('place'))
            expense.trip = shiftmodel.objects.get(pk=request.POST.get('trip'))
            expense.driver = drivermodel.objects.get(driver=request.POST.get('driver'))
            expense.dat = request.POST.get('dat')
            expense.feet = request.POST.get('feet')
            expense.hire = request.POST.get('hire')
            expense.halt = request.POST.get('halt')
            expense.dis = request.POST.get('dis')
            expense.bno = request.POST.get('bno')
            expense.disqnt = request.POST.get('disqnt')
            expense.updated_by = request.user
            expense.cont = request.POST.get('cont')
            if expense.feet == '20R' or expense.feet == '40R' or expense.feet == '2X20R':
                expense.batta = int(expense.hire) * 0.18
            elif expense.feet == '20' or expense.feet == '40' or expense.feet == '2X20':
                try:
                    if expense.feet == '2X20':
                        bata = battaregmodel.objects.get(place=expense.place, feet='40')
                        expense.batta = bata.batta
                    else:
                        bata = battaregmodel.objects.get(place=expense.place, feet=expense.feet)
                        expense.batta = bata.batta
                except battaregmodel.DoesNotExist:
                    expense.batta = 0
                    messages.warning(request, "Batta rate not found for this place and feet")
            else:
                expense.batta=request.POST.get('batta')
            expense.save()
            messages.success(request, "Trip updated successfully.")

            if bill.exists():
                bill.update(hire=expense.hire)
            if batta.exists():
                batta.update(driver=expense.driver)

            return redirect('triporderupdate')
        except Exception as e:
            messages.error(request, str(e))
    return render(request, 'staff/ISA/trip/isaup.html', {'code': expense, 'party': party, 'place': place,
                                                         'type': type,'driver':driver})

@transaction.atomic
def isadel(request, id):
    try:
        trip = tripmodel.objects.get(pk=id)
    except tripmodel.DoesNotExist:
        messages.error(request, "Trip not found.")
        return redirect('triporderupdate')

    try:
        batta = battamodel.objects.get(pk=trip.battaid)
        batta.deleted_by = request.user
        batta.delete()
    except battamodel.DoesNotExist:
        pass

    try:
        exp = expensemodel.objects.get(pk=trip.expid)
        exp.deleted_by = request.user
        exp.delete()
    except expensemodel.DoesNotExist:
        pass

    trip.deleted_by = request.user
    trip.delete()
    messages.success(request, "Trip deleted successfully.")
    return redirect('triporderupdate')

def triphistory(request):
    if 'username' in request.session:
        isa_trucks = truckmodel.objects.all()
        if 'ser1' in request.POST:
            try:
                frdate = request.POST.get('frdate')
                todate = request.POST.get('todate')
                code = request.POST.get('code')

                if(not frdate or frdate == '') and (not todate or todate == '') and (
                            code == '' or code == '----------------------------------Select---------------------------------------'):
                    messages.error(request, "Select any filter")
                    return render(request, 'admin/history/triphistory.html', {'code': isa_trucks})

                filter_kwargs = {}

                if code and code != '----------------------------------Select---------------------------------------':
                    filter_kwargs['code'] = code

                if frdate and todate:
                    filter_kwargs['dat__range'] = [frdate, todate]

                trips = tripmodel.objects.filter(**filter_kwargs).order_by('-dat')
                if trips.exists():
                    combined_data = []
                    diesel = 0
                    hire = 0
                    for trip in trips:
                        bill_data = billmodel.objects.filter(bno=trip.bno)
                        trans_data = truckmodel.objects.filter(code=trip.code).first()
                        if trans_data:
                            diesel += trip.dis
                            hire += trip.hire
                        combined_data.append({'trip': trip, 'bill': bill_data, 'trans': trans_data})

                    return render(request, 'admin/history/triphistory.html', {'ser1': combined_data, 'diesel': diesel, 'hire': hire, 'code': isa_trucks})
                else:
                    messages.error(request, 'No trips found.')
                    return render(request, 'admin/history/triphistory.html', {'code': isa_trucks})
            except tripmodel.DoesNotExist:
                messages.error(request, 'No trips found.')
                return render(request, 'admin/history/triphistory.html', {'code': isa_trucks})

        if 'trip_data' in request.session:
            combined_data = request.session['trip_data']
            diesel = request.session['diesel']
            hire = request.session['hire']
            return render(request, 'admin/history/triphistory.html', {'ser1': combined_data, 'diesel': diesel,
                                                                      'hire': hire, 'code': isa_trucks})
        return render(request, 'admin/history/triphistory.html', {'code': isa_trucks})
    else:
        return redirect('login')

def triporderhistory(request, id):
    if 'username' in request.session:
        try:
            trip = tripmodel.objects.get(id=id)
            history = trip.history.all()
            history.update(created_at=timezone.localtime(trip.created_at), updated_at=timezone.localtime(trip.updated_at))
            return render(request, 'admin/history/triporderhistory.html', {'history': history})
        except tripmodel.DoesNotExist:
            messages.error(request, 'Trip not found')
            return render(request, 'admin/history/triporderhistory.html')
        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'admin/history/triporderhistory.html')
    else:
        return redirect('login')

def deltriphistory(request):
    """Handles trip history deletion."""
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.all()

    if request.method == 'POST' and 'ser1' in request.POST:
        try:
            frdate = request.POST.get('frdate')
            todate = request.POST.get('todate')
            code = request.POST.get('code')

            if not (frdate or todate or code):
                messages.error(request, "Select any filter")
                return render(request, 'admin/history/deltriphistory.html', {'code': isa_trucks})

            filter_kwargs = {'history_type__in': ['-']}
            if code and code != '----------------------------------Select---------------------------------------':
                filter_kwargs['code'] = code
            if frdate and todate:
                filter_kwargs['dat__range'] = [frdate, todate]

            deleted_trips = tripmodel.history.filter(**filter_kwargs).order_by('-dat')
            if deleted_trips.exists():
                return render(request, 'admin/history/deltriphistory.html', {'ser1': deleted_trips, 'code': isa_trucks})
            else:
                messages.error(request, 'No trips found.')
                return render(request, 'admin/history/deltriphistory.html', {'code': isa_trucks})
        except tripmodel.DoesNotExist:
            messages.error(request, 'No trips found.')
            return render(request, 'admin/history/deltriphistory.html', {'code': isa_trucks})
    return render(request, 'admin/history/deltriphistory.html', {'code': isa_trucks})

def deltriporderhistory(request, id):
    """Handles trip order history deletion."""
    if 'username' not in request.session:
        return redirect('login')

    try:
        history = tripmodel.history.filter(id=id)
        if not history.exists():
            messages.error(request, 'Trip not found')
            return render(request, 'admin/history/deltriporderhistory.html')
        return render(request, 'admin/history/deltriporderhistory.html', {'history': history})
    except Exception as e:
        messages.error(request, str(e))
        return render(request, 'admin/history/deltriporderhistory.html')

def vehicle(request):
    if 'username' not in request.session:
        return redirect('login')

    today = timezone.localdate()

    try:
        trip = tripmodel.objects.filter(code__trans__in=['ISA','THOLADAN','ISA TRANS']).filter(~Q(dat=today))

        if trip.exists():
            latest_trip = trip.latest('dat')
            latest_day = latest_trip.dat
            diff = (today - latest_day).days

            if diff >= 1:
                day = today - timezone.timedelta(days=diff)
                day = day.strftime('%Y-%m-%d')

                current_trips = tripmodel.objects.filter(dat=today).exclude(place__in=['NO WORK', 'WORKSHOP'])
                previous_trips = tripmodel.objects.filter(dat=day, place__in=['NO WORK', 'WORKSHOP']).exclude(code__in=[t.code for t in current_trips]).order_by('dat')

                return render(request, 'staff/ISA/trip/vehicle.html', {'trucks': previous_trips})
            else:
                return render(request, 'staff/ISA/trip/vehicle.html')
        else:
            return render(request, 'staff/ISA/trip/vehicle.html')

    except Exception as e:
        print(f"An error occurred: {e}")
        return render(request, 'staff/ISA/trip/vehicle.html')

def outtriporder(request):
    if 'username' not in request.session:
        return redirect('login')
    x = placemodel.objects.filter(checks=True)
    y = shiftmodel.objects.filter(checks=True)
    z = partymodel.objects.filter(checks=True)
    tran = transportermodel.objects.exclude(trans__in=['ISA', 'THOLADAN', 'ISA TRANS'])
    a = truckmodel.objects.exclude(trans__in=['ISA', 'THOLADAN', 'ISA TRANS'])
    today = date.today().strftime('%Y-%m-%d')
    if request.method == 'POST':
        truck_no = request.POST.get('code')
        feet = request.POST.get('feet')
        driver = drivermodel.objects.get(driver=request.POST.get('driver'))
        tran_name = request.POST.get('trans')
        try:
            trans = transportermodel.objects.get(trans=tran_name)
        except transportermodel.DoesNotExist:
            trans = transportermodel.objects.create(trans=tran_name)
        try:
            truck = truckmodel.objects.get(truckno=truck_no)
            driver = truck.driver
        except truckmodel.DoesNotExist:
            truck = truckmodel.objects.create(truckno=truck_no, code=truck_no, feet=feet, driver=driver, trans=trans)
        dat = request.POST.get('dat')
        dis = request.POST.get('dis') or 0
        hire = request.POST.get('hire') or 0
        halt = request.POST.get('halt') or 0
        disqnt = request.POST.get('disqnt') or 0
        buy = request.POST.get('buy') or 0
        outadv = request.POST.get('outadv') or 0
        partyid = request.POST.get('party')
        try:
            party = partymodel.objects.get(party=partyid)
        except partymodel.DoesNotExist:
            messages.error(request, "Party not found")
            return redirect('outtriporder')
        placeid = request.POST.get('place')
        try:
            place = placemodel.objects.get(place=placeid)
        except placemodel.DoesNotExist:
            messages.error(request, "Place not found")
            return redirect('outtriporder')
        tripid = request.POST.get('trip')
        try:
            trip = shiftmodel.objects.filter(checks=True).get(trip=tripid)
        except shiftmodel.DoesNotExist:
            messages.error(request, "Trip not found")
            return redirect('outtriporder')
        created_by = request.user
        cont = request.POST.get('cont') or 'NIL'
        bno = request.POST.get('bno') or 0
        tripmodel.objects.create(
            code=truck,
            driver=driver,
            dat=dat,
            party=party,
            place=place,
            trip=trip,
            dis=dis,
            bno=bno,
            disqnt=disqnt,
            buy=buy,
            created_by=created_by,
            hire=hire,
            halt=halt,
            cont=cont,
            outadv=outadv
        )
        messages.success(request, "Trip order created successfully.")
        return redirect('outtriporder')
    else:
        return render(request, 'staff/out/outtriporder.html', {"d": x, "d1": y, "d2": z, "d3": a, 'trans': tran, 'today': today})

def outtriporderupdate(request):
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.exclude(trans__in=['ISA','THOLADAN','ISA TRANS'])
    filter_params = request.session.get('filter_params')

    if request.method == 'POST':
        try:
            frdate = request.POST.get('frdate')
            todate = request.POST.get('todate')
            code = request.POST.get('code')
            request.session['filter_params'] = {
                'frdate': frdate,
                'todate': todate,
                'code': code,
            }

            if (not frdate or frdate == '') and (not todate or todate == '') and (code == '' or code == '----------------------------------Select---------------------------------------'):
                messages.error(request, "Select any filter")
                return render(request, 'staff/out/outtriporderupdate.html', {'code': isa_trucks, 'filter_params': request.session.get('filter_params')})

            filter_kwargs = {}
            if code and code != '----------------------------------Select---------------------------------------':
                filter_kwargs['code'] = code
            if frdate and todate:
                filter_kwargs['dat__range'] = [frdate, todate]

            trips = tripmodel.objects.filter(**filter_kwargs).order_by('dat')

            if 'ser1' in request.POST:
                if trips.exists():
                    combined_data = []
                    diesel = 0
                    hire = 0
                    halt = 0
                    for trip in trips:
                        trans_data = truckmodel.objects.filter(code=trip.code).exclude(trans__in=['ISA','THOLADAN','ISA TRANS']).first()
                        if trans_data:
                            diesel += trip.dis
                            halt += trip.halt
                            hire += trip.hire
                            combined_data.append({'trip': trip})
                    return render(request, 'staff/out/outtriporderupdate.html', {'ser1': combined_data, 'diesel': diesel, 'hire': hire, 'code': isa_trucks, 'halt': halt, 'filter_params': request.session.get('filter_params')})
                else:
                    messages.error(request, 'No trips found.')
                    return render(request, 'staff/out/outtriporderupdate.html', {'code': isa_trucks, 'filter_params': request.session.get('filter_params')})


            elif 'export' in request.POST:

                if trips.exists():
                    diesel = 0
                    hire = 0
                    halt = 0
                    combined = []
                    for trip in trips:
                        trans_data = truckmodel.objects.filter(code=trip.code).exclude(trans__in=['ISA', 'THOLADAN', 'ISA TRANS']).first()
                        if trans_data:
                            diesel += trip.dis
                            halt += trip.halt
                            hire += trip.hire
                        combined.append({'trip': trip})
                    total = hire + halt
                    html = render_to_string('staff/out/outtripreport.html',
                                            {'trip': combined, 'diesel': diesel, 'hire': hire, 'total': total,'halt': halt})
                    wb = Workbook()
                    ws = wb.active
                    soup = BeautifulSoup(html, 'html.parser')
                    table = soup.find('table')
                    rows = table.find_all('tr')
                    row_idx = 1
                    for row in rows:
                        cells = row.find_all('td')
                        for i, cell in enumerate(cells):
                            value = cell.text
                            if value.replace('.', '', 1).replace('-', '', 1).isdigit():
                                ws.cell(row=row_idx, column=i + 1).value = float(value)
                                ws.cell(row=row_idx, column=i + 1).number_format = '#,##0.00;[Red]-#,##0.00'
                            else:
                                ws.cell(row=row_idx, column=i + 1).value = value
                            if row_idx == 1:
                                ws.cell(row=row_idx, column=i + 1).font = Font(bold=True)
                        row_idx += 1
                    wb.save('tripreport.xlsx')
                    response = HttpResponse(
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'filename="tripreport.xlsx"'
                    wb.save(response)
                    return response

        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'staff/out/outtriporderupdate.html', {'code': isa_trucks, 'filter_params': request.session.get('filter_params')})

    elif filter_params:
        try:
            frdate = filter_params.get('frdate')
            todate = filter_params.get('todate')
            code = filter_params.get('code')

            filter_kwargs = {}
            if code and code != '----------------------------------Select---------------------------------------':
                filter_kwargs['code'] = code
            if frdate and todate:
                filter_kwargs['dat__range'] = [frdate, todate]

            trips = tripmodel.objects.filter(**filter_kwargs).order_by('dat')

            if trips.exists():
                combined_data = []
                diesel = 0
                hire = 0
                halt = 0
                for trip in trips:
                    trans_data = truckmodel.objects.filter(code=trip.code).exclude(trans__in=['ISA','THOLADAN','ISA TRANS']).first()
                    if trans_data:
                        diesel += trip.dis
                        halt += trip.halt
                        hire += trip.hire
                        combined_data.append({'trip': trip})
                return render(request, 'staff/out/outtriporderupdate.html', {'ser1': combined_data, 'diesel': diesel, 'hire': hire, 'code': isa_trucks, 'halt': halt, 'filter_params': filter_params})
            else:
                messages.error(request, 'No trips found.')
                return render(request, 'staff/out/outtriporderupdate.html', {'code': isa_trucks, 'filter_params': filter_params})

        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'staff/out/outtriporderupdate.html', {'code': isa_trucks, 'filter_params': filter_params})

    return render(request, 'staff/out/outtriporderupdate.html', {'code': isa_trucks})

def outup(request,id):
    if 'username' not in request.session:
        return redirect('login')

    party = partymodel.objects.all()
    place = placemodel.objects.all()
    type = shiftmodel.objects.filter(checks=True)

    try:
        expense = tripmodel.objects.get(pk=id)
    except tripmodel.DoesNotExist:
        messages.error(request, "Expense not found.")
        return redirect('outtriporderupdate')

    if request.method == 'POST':
        try:
            expense.party = partymodel.objects.get(pk=request.POST.get('party'))
            expense.place = placemodel.objects.get(pk=request.POST.get('place'))
            expense.trip = shiftmodel.objects.get(pk=request.POST.get('trip'))
            expense.hire = request.POST.get('hire')
            expense.halt = request.POST.get('halt')
            expense.dis = request.POST.get('dis')
            expense.bno = request.POST.get('bno')
            expense.updated_by = request.user
            expense.cont = request.POST.get('cont')
            expense.buy = request.POST.get('buy')
            expense.outadv = request.POST.get('outadv')
            expense.save()
            messages.success(request, "Expense updated successfully.")
            return redirect('outtriporderupdate')
        except Exception as e:
            messages.error(request, str(e))
            return render(request, 'staff/out/outup.html', {'code': expense, 'party': party, 'place': place, 'type': type})
    else:
        return render(request, 'staff/out/outup.html', {'code': expense, 'party': party, 'place': place, 'type': type})

@transaction.atomic
def outdel(request, id):
    try:
        trip = get_object_or_404(tripmodel, pk=id)
        trip.delete()
        messages.success(request, "Trip deleted successfully.")
        return redirect('outtriporderupdate')
    except Exception as e:
        messages.error(request, str(e))
        return redirect('outtriporderupdate')

def handle_form_submission(request):
    """ Handles the form submission for the bill page.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: The HTTP response object.
    """
    if request.method == 'POST' and 'ser' in request.POST:
        start_date = request.POST.get('dat')
        end_date = request.POST.get('todate')
        party = request.POST.getlist('party')
        place = request.POST.getlist('place')

        # Check if any filters are selected
        if not party and not place and not start_date and not end_date:
            messages.error(request, "Select any filter")
            return render(request, 'staff/ISA/bill/bill.html', {'places': placemodel.objects.filter(checks=True), 'parties': partymodel.objects.filter(checks=True)})

        # Filter trips based on the selected criteria
        filter_kwargs = {
            'bill': False,
        }
        if party:
            filter_kwargs['party__in'] = party
        if place:
            filter_kwargs['place__in'] = place
        if start_date and end_date:
            filter_kwargs['dat__range'] = [start_date, end_date]

        trips = tripmodel.objects.filter(**filter_kwargs).order_by('dat').exclude(place__in=['NO WORK', 'WORKSHOP'])

        # Render the bill page with the filtered trips
        if trips.exists():
            combined = []
            for trip in trips:
                trans_data = truckmodel.objects.filter(code=trip.code)
                if trans_data:
                    combined.append({'trip': trip})
            return render(request, 'staff/ISA/bill/bill.html', {'trips': combined, 'places': placemodel.objects.filter(checks=True), 'parties': partymodel.objects.filter(checks=True)})
        else:
            messages.error(request, 'No trips found.')
            return render(request, 'staff/ISA/bill/bill.html', {'places': placemodel.objects.filter(checks=True), 'parties': partymodel.objects.filter(checks=True)})

def generate_bill(request):
    """
    Generates the bill for the selected trips.
    """
    if request.method == 'POST' and 'gen' in request.POST:
        id_list = request.POST.getlist('boxes')
        try:
            for trip_code in id_list:
                trips = tripmodel.objects.filter(pk=trip_code)
                trips.update(checked=True)
            return redirect('billing')
        except tripmodel.DoesNotExist:
            messages.error(request, "No trips found.")
            return redirect('bill')

def bill(request):
    """Renders the bill page."""
    if 'username' not in request.session:
        return redirect('login')

    parties = partymodel.objects.filter(checks=True)
    places = placemodel.objects.filter(checks=True)

    if request.method == 'POST':
        try:
            if 'ser' in request.POST:
                return handle_form_submission(request)
            elif 'gen' in request.POST:
                return generate_bill(request)
            else:
                messages.error(request, "Invalid request.")
        except Exception as e:
            messages.error(request, str(e))

    return render(request, 'staff/ISA/bill/bill.html', {'places': places, 'parties': parties})

def billing(request):
    """Renders the billing page."""
    if 'username' not in request.session:
        return redirect('login')

    try:
        trips = tripmodel.objects.filter(checked=True)
        return render(request, 'staff/ISA/bill/billing.html', {'trips': trips})
    except Exception as e:
        messages.error(request, str(e))
        return redirect('login')

def delrow(request, id):
    """Deletes a row by setting checked to False."""
    try:
        trip = tripmodel.objects.get(id=id)
        trip.checked = False
        trip.save()
        return redirect('billing')
    except tripmodel.DoesNotExist:
        messages.error(request, "Trip not found.")
        return redirect('billing')

def billdetails(request):
    """Renders the bill details page."""
    if 'username' not in request.session:
        return redirect('login')

    try:
        hire = 0
        halt = 0
        dis = 0
        haltqnt = 0
        hireqnt = 0
        cont = []
        last_bill = billmodel.objects.all().order_by('bno').last()
        if last_bill:
            bno = last_bill.bno + 1
        else:
            bno = 1

        trips = tripmodel.objects.filter(checked=True)
        for trip in trips:
            halt = trip.halt
            dis += trip.dis
            if halt > 0:
                haltqnt += 1
            if trip.hire > 0:
                hireqnt += 1
            if trip.cont not in ['NIL','SAME']:
                cont.append(trip.cont)

        for trip in trips:
            hire = trip.hire
            break

        bdate = timezone.localdate()

        if request.method == 'POST':
            if 'down' in request.POST:
                try:
                    bno = request.POST.get('bno')
                    diesel = request.POST.get('diesel')
                    hire = int(request.POST.get('hire') or 0)
                    hireqnt = int(request.POST.get('hireqnt') or 0)
                    toll = int(request.POST.get('toll') or 0)
                    tollqnt = int(request.POST.get('tollqnt') or 0)
                    unload = int(request.POST.get('unload') or 0)
                    unloadqnt = int(request.POST.get('unloadqnt') or 0)
                    enblock = int(request.POST.get('enblock') or 0)
                    enblockqnt = int(request.POST.get('enblockqnt') or 0)
                    shift = int(request.POST.get('shift') or 0)
                    shiftqnt = int(request.POST.get('shiftqnt') or 0)
                    weigh = int(request.POST.get('weigh') or 0)
                    halt = int(request.POST.get('halt') or 0)
                    weighqnt = int(request.POST.get('weighqnt') or 0)
                    haltqnt = int(request.POST.get('haltqnt') or 0)
                    total = (hire * hireqnt) + (toll * tollqnt) + (unload * unloadqnt) + (enblock * enblockqnt) + (
                                shift * shiftqnt) + (weigh * weighqnt) + (halt * haltqnt) - float(diesel or 0)
                    bill = billmodel.objects.create(
                        bno=bno,
                        bdate=bdate,
                        diesel=diesel,
                        hire=hire,
                        hireqnt=hireqnt,
                        toll=toll,
                        tollqnt=tollqnt,
                        unload=unload,
                        unloadqnt=unloadqnt,
                        enblock=enblock,
                        enblockqnt=enblockqnt,
                        shift=shift,
                        shiftqnt=shiftqnt,
                        weigh=weigh,
                        halt=halt,
                        weighqnt=weighqnt,
                        haltqnt=haltqnt,
                        total=total,
                        created_by=request.user
                    )
                    bill.save()
                    tripmodel.objects.filter(checked=True).update(bno=bno, updated_by=request.user)
                    messages.success(request, "Bill saved successfully.")
                    return redirect('billpdf')
                except Exception as e:
                    messages.error(request, f"Error saving bill: {e}")
                    print(f"Error saving bill: {e}")
                    return redirect('billdetails')
        return render(request, 'staff/ISA/bill/billdetails.html', {
            'hireqnt': hireqnt,
            'hire': hire,
            'halt': halt,
            'haltqnt': haltqnt,
            'dis': dis,
            'cont': cont,
            'bno': bno,
            'bdate': bdate
        })
    except Exception as e:
        messages.error(request, f"Error rendering bill details: {e}")
        print(f"Error rendering bill details: {e}")
        return redirect('login')

def billpdf(request):
    trip=tripmodel.objects.filter(checked=True)
    conts=[]
    code=[]
    for i in trip:
        bills=i.dat

    for i in trip:
        if i.cont not in ['NIL','SAME']:
            cont_values = i.cont.split(',')  # split on comma
            conts.extend(cont_values)  # append each value to conts
            code.append(i.code)
    word_count = len(conts)
    for i in trip:
        place=i.place
        party=i.party
        feet=i.code.feet
        date=i.dat
        add=i.party.add
        break
    bill = billmodel.objects.filter(bno=trip.first().bno)
    ttoll=tweigh=tenblock=tunload=thalt=tshift=total=thire=0
    words=''
    day=0
    for i in bill:
        thire += i.hireqnt * i.hire
        ttoll += i.tollqnt * i.toll
        tweigh += i.weighqnt * i.weigh
        tenblock += i.enblockqnt * i.enblock
        tunload += i.unloadqnt * i.unload
        thalt += i.haltqnt * i.halt
        tshift += i.shiftqnt * i.shift
        total += i.total
        words += num2words.num2words(total)
    bill.update(total=total)
    current_year = datetime.now().year
    if datetime.now().month == 12 and datetime.now().day > 31:
        next_academic_year = f"{current_year+1}-{str(current_year + 2)[-2:]}"
    else:
        next_academic_year = f"{current_year}-{str(current_year + 1)[-2:]}"
    qr_path = 'myproject/static/img/tholadan/tholadqr.jpg'
    qr = cv2.imread(qr_path)
    resized_img = cv2.resize(qr, (150, 150))
    success, buffer = cv2.imencode('.png', resized_img)
    qr_data = buffer.tobytes()
    qr_image = base64.b64encode(qr_data).decode('utf-8')
    template = get_template('staff/ISA/bill/ex.html')

    context = {'qr': qr_image,'trip':trip,'bill':bill,'place':place,'party':party,'feet':feet,'add':add,'ttoll': ttoll,
               'thire': thire, 'words': words,'tweigh': tweigh, 'tenblock': tenblock, 'tunload': tunload,'thalt': thalt,
               'tshift': tshift,'total': total,'count': word_count,'day':day,'dat':date,'billdate':bills,
               'conts':conts,'code':code,'next_academic_year': next_academic_year}
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline filename="bill.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    else:
        tripmodel.objects.filter(checked=True).update(checked=False,bill=True,updated_by=request.user)
    return response

def viewbill(request):
    """Renders the view bill page."""
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
    parties = partymodel.objects.all()
    places = placemodel.objects.all()

    if request.method == 'POST' and 'ser1' in request.POST:
        try:
            frdate = request.POST.get('frdate')
            todate = request.POST.get('todate')
            code = request.POST.get('code')
            placeid = request.POST.get('place')
            partyid = request.POST.get('party')
            bno = request.POST.get('bno')
            container = request.POST.get('cont')
            filters = {}
            if container:
                container_list = [c.strip() for c in container.split(',')]
                trip_filters = Q()
                for c in container_list:
                    trip_filters |= Q(cont__icontains=c)
                trip_bnos = tripmodel.objects.filter(trip_filters).values_list('bno', flat=True)
                bills = billmodel.objects.filter(bno__in=trip_bnos)
            else:
                bills = billmodel.objects.all()

            if frdate and todate:
                bills = bills.filter(bdate__range=[frdate, todate])
            if bno:
                bills = bills.filter(bno=bno)

            bills = bills.order_by('bno')
            if not bills and bno:
                messages.error(request, 'Bill not found.')
                return render(request, 'staff/ISA/bill/viewbill.html', {
                    'code': isa_trucks,
                    'parties': parties,
                    'places': places
                })

            combined_data = []
            total = 0
            for bill in bills:
                total += bill.total
                combined_data.append({'bill': bill})

            return render(request, 'staff/ISA/bill/viewbill.html', {
                'bill': combined_data,
                'total': total,
                'code': isa_trucks,
                'parties': parties,
                'places': places
            })
        except (tripmodel.DoesNotExist, placemodel.DoesNotExist, partymodel.DoesNotExist) as e:
            messages.error(request, f'No trips found: {e}')
            return render(request, 'staff/ISA/bill/viewbill.html', {
                'code': isa_trucks,
                'parties': parties,
                'places': places
            })

    return render(request, 'staff/ISA/bill/viewbill.html', {
        'code': isa_trucks,
        'parties': parties,
        'places': places
    })

def billhistory(request):
    if 'username' in request.session:
        isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
        parties = partymodel.objects.all()
        places = placemodel.objects.all()
        if 'ser1' in request.POST:
            try:
                frdate = request.POST.get('frdate')
                todate = request.POST.get('todate')
                bno = request.POST.get('bno')
                if not bno and not frdate and not todate:
                    messages.error(request, "Select any filter")
                    return render(request, 'admin/history/billhistory.html', {
                        'code': isa_trucks,
                        'parties': parties,
                        'places': places
                    })
                filters = {}
                if frdate and todate:
                    filters['bdate__range'] = [frdate, todate]
                if bno:
                    filters['bno'] = bno
                bills = billmodel.objects.filter(**filters).order_by('bno')
                combined_data = []
                if bills.exists():
                    for bill in bills:
                        combined_data.append({'bill': bill})
                    return render(request, 'admin/history/billhistory.html', {
                    'bill': combined_data,
                    'code': isa_trucks,
                    'parties': parties,
                    'places': places,
                })
                else:
                    messages.error(request, 'Bill not found.')
            except (tripmodel.DoesNotExist, placemodel.DoesNotExist, partymodel.DoesNotExist,billmodel.DoesNotExist):
                messages.error(request, 'No trips found.')
                return render(request, 'admin/history/billhistory.html', {
                    'code': isa_trucks,
                    'parties': parties,
                    'places': places
                })
        return render(request, 'admin/history/billhistory.html', {
            'code': isa_trucks,
            'parties': parties,
            'places': places
        })
    else:
        return redirect('login')

def billorderhistory(request, id):
    if 'username' in request.session:
        try:
            trip = billmodel.objects.get(pk=id)
            history = trip.history.all()
            history.update(created_at=timezone.localtime(trip.created_at), updated_at=timezone.localtime(trip.updated_at))
            return render(request, 'admin/history/billorderhistory.html', {'history': history})
        except billmodel.DoesNotExist:
            request.session['error'] = 'Trip not found'
            return render(request, 'admin/history/billorderhistory.html', {'error': 'Bill not found'})
        except Exception as e:
            request.session['error'] = str(e)
            return render(request, 'admin/history/billorderhistory.html', {'error': str(e)})
    else:
        return redirect('login')

def billupdate(request,bno):
    if 'username' in request.session:
        billd = request.POST.get('hdate')
        trips = tripmodel.objects.filter(bno=bno)
        hbill=''
        bdates = timezone.localdate()
        print(bdates)
        for i in trips:
            if i.halt>0:
                hbill = i.dat
        try:
            bill = billmodel.objects.get(pk=bno)
        except billmodel.DoesNotExist:
            messages.error(request, "Bill not found.")
            return redirect('billupdate')

        if request.method == 'POST':
            if 'update' in request.POST:
                trip = tripmodel.objects.filter(bno=bno).exclude(place='SAME')
                bill.bdate = request.POST.get('bdate')
                bill.diesel = request.POST.get('diesel')
                bill.hire = request.POST.get('hire')
                trip.hire = request.POST.get('hire')
                bill.hireqnt = request.POST.get('hireqnt')
                bill.halt = request.POST.get('halt')
                bill.haltqnt = request.POST.get('haltqnt')
                bill.weigh = request.POST.get('weigh')
                bill.weighqnt = request.POST.get('weighqnt')
                bill.unload = request.POST.get('unload')
                bill.unloadqnt = request.POST.get('unloadqnt')
                bill.toll = request.POST.get('toll')
                bill.tollqnt = request.POST.get('tollqnt')
                bill.enblock = request.POST.get('enblock')
                bill.enblockqnt = request.POST.get('enblockqnt')
                bill.shift = request.POST.get('shift')
                bill.shiftqnt = request.POST.get('shiftqnt')
                bill.updated_by = request.user
                bill.save()
                trip.update(hire=bill.hire)
                return redirect('viewbill')

            if 'down' in request.POST:
                bill = billmodel.objects.filter(bno=bno)
                trip = tripmodel.objects.filter(bno=bno)
                conts = []
                code = []
                for i in trip:
                    hbill=i.dat
                    if i.cont not in ['NIL','SAME']:
                        cont_values = i.cont.split(',')  # split on comma
                        conts.extend(cont_values)  # append each value to conts
                        code.append(i.code)
                if billd:
                    bills =billd
                else:
                    bills = hbill
                print(bills)
                word_count = len(conts)
                for i in trip:
                    place = i.place
                    party = i.party
                    feet = i.code.feet
                    date = i.dat
                    add = i.party.add
                    break
                ttoll = tweigh = tenblock = tunload = thalt = tshift = total = thire = 0
                words = ''
                day = 0
                for i in bill:
                    thire += i.hireqnt * i.hire
                    ttoll += i.tollqnt * i.toll
                    tweigh += i.weighqnt * i.weigh
                    tenblock += i.enblockqnt * i.enblock
                    tunload += i.unloadqnt * i.unload
                    thalt += i.haltqnt * i.halt
                    tshift += i.shiftqnt * i.shift
                    total += i.total
                    words += num2words.num2words(total)
                print(conts)
                print('container:', word_count)
                qr_path = 'myproject/static/img/tholadan/tholadqr.jpg'
                qr = cv2.imread(qr_path)
                resized_img = cv2.resize(qr, (150, 150))
                success, buffer = cv2.imencode('.png', resized_img)
                qr_data = buffer.tobytes()
                qr_image = base64.b64encode(qr_data).decode('utf-8')
                template = get_template('staff/ISA/bill/ex.html')
                current_year = datetime.now().year
                bill.update(total=total)
                if datetime.now().month == 12 and datetime.now().day > 31:
                    next_academic_year = f"{current_year+1}-{str(current_year + 2)[-2:]}"
                else:
                    next_academic_year = f"{current_year}-{str(current_year + 1)[-2:]}"
                context = {'qr': qr_image, 'trip': trip, 'bill': bill, 'ttoll': ttoll,
                           'thire': thire, 'words': words, 'tweigh': tweigh, 'tenblock': tenblock, 'tunload': tunload,
                           'thalt': thalt,'bdate':bdates,
                           'tshift': tshift, 'total': total, 'count': word_count, 'day': day,
                           'next_academic_year': next_academic_year,
                           'billdate': bills,
                           'conts': conts, 'code': code, 'place': place, 'party': party, 'feet': feet,
                           'add': add, 'dat': date}
                html = template.render(context)

                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = 'inline; filename="bill.pdf"'
                pisa_status = pisa.CreatePDF(html, dest=response)

                if pisa_status.err:
                    return HttpResponse('We had some errors <pre>' + html + '</pre>')
                return response
        return render(request,'staff/ISA/bill/billupdate.html',{'bill':bill,'billdate':hbill,'bdate':bdates})
    else:
        return redirect('login')

def dailyupdate(request):
    """Renders the daily update page."""
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])

    if request.method == 'POST' and 'ser1' in request.POST:
        try:
            frdate = request.POST.get('frdate')
            todate = request.POST.get('todate')
            code = request.POST.get('code')

            if (not frdate or frdate == '') and (not todate or todate == '') and (code == '' or code == '----------------------------------Select---------------------------------------'):
                messages.error(request, 'Select any filters')
                return render(request, 'admin/ISA/dailyupdate.html', {'code': isa_trucks})

            filter_kwargs = {}
            if code and code != '----------------------------------Select---------------------------------------':
                filter_kwargs['code'] = code
            if frdate and todate:
                filter_kwargs['dat__range'] = [frdate, todate]

            trips = tripmodel.objects.filter(**filter_kwargs).order_by('dat')

            if trips.exists():
                combined_data = []
                diesel = 0
                hire = 0
                halt = 0

                for trip in trips:
                    trans_data = truckmodel.objects.filter(code=trip.code, trans__in=['ISA','THOLADAN','ISA TRANS']).first()
                    if trans_data:
                        diesel += trip.dis
                        halt += trip.halt
                        hire += trip.hire
                    combined_data.append({'trip': trip})

                return render(request, 'admin/ISA/dailyupdate.html', {'trip': combined_data, 'halt': halt, 'diesel': diesel, 'hire': hire, 'code': isa_trucks})
            else:
                messages.error(request, 'No trips found.')
                return render(request, 'admin/ISA/dailyupdate.html', {'code': isa_trucks})
        except tripmodel.DoesNotExist:
            messages.error(request, 'No trips found.')
            return render(request, 'admin/ISA/dailyupdate.html', {'code': isa_trucks})

    return render(request, 'admin/ISA/dailyupdate.html', {'code': isa_trucks})

def stafftripview(request):
    """Renders the staff trip view page."""
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])

    if request.method == 'POST':
        diesel = 0
        hire = 0
        halt = 0
        frdate = request.POST.get('frdate')
        todate = request.POST.get('todate')
        code = request.POST.get('code')

        if (not frdate or frdate == '') and (not todate or todate == '') and (code == '' or code == '----------------------------------Select---------------------------------------'):
            messages.error(request, 'Select any filters')
            return render(request, 'staff/reports/stafftripview.html', {'code': isa_trucks})

        filter_kwargs = {}
        if code and code != '----------------------------------Select---------------------------------------':
            filter_kwargs['code'] = code
        if frdate and todate:
            filter_kwargs['dat__range'] = [frdate, todate]

        trips = tripmodel.objects.filter(**filter_kwargs).order_by('dat')

        if 'ser1' in request.POST:
            if trips.exists():
                combined = []
                for trip in trips:
                    trans_data = truckmodel.objects.filter(code=trip.code,trans__in=['ISA','THOLADAN','ISA TRANS']).first()
                    if trans_data:
                        diesel += trip.dis
                        halt += trip.halt
                        hire += trip.hire
                    combined.append({'trip': trip})
                return render(request, 'staff/reports/stafftripview.html', {'trip': combined, 'halt': halt, 'diesel': diesel, 'hire': hire, 'code': isa_trucks})
            else:
                messages.error(request, 'Trip not found')

        elif 'export' in request.POST:
            if trips.exists():
                combined = []
                for trip in trips:
                    trans_data = truckmodel.objects.filter(code=trip.code, trans__in=['ISA','THOLADAN','ISA TRANS']).first()
                    if trans_data:
                        diesel += trip.dis
                        halt += trip.halt
                        hire += trip.hire
                    combined.append({'trip': trip})

                total = hire + halt
                html = render_to_string('staff/reports/tripreport.html', {'trip': combined, 'diesel': diesel, 'hire': hire, 'total': total, 'halt': halt})

                wb = Workbook()
                ws = wb.active
                soup = BeautifulSoup(html, 'html.parser')
                table = soup.find('table')
                rows = table.find_all('tr')

                row_idx = 1
                for row in rows:
                    cells = row.find_all('td')
                    for i, cell in enumerate(cells):
                        value = cell.text
                        if value.replace('.', '', 1).replace('-', '', 1).isdigit():
                            ws.cell(row=row_idx, column=i + 1).value = float(value)
                            ws.cell(row=row_idx, column=i + 1).number_format = '#,##0.00;[Red]-#,##0.00'
                        else:
                            ws.cell(row=row_idx, column=i + 1).value = value
                        if row_idx == 1:
                            ws.cell(row=row_idx, column=i + 1).font = Font(bold=True)
                    row_idx += 1

                wb.save('tripreport.xlsx')
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'filename="tripreport.xlsx"'
                wb.save(response)
                return response

    return render(request, 'staff/reports/stafftripview.html', {'code': isa_trucks})

def outdailyupdate(request):
    """Renders the out daily update page."""
    if 'username' not in request.session:
        return redirect('login')

    x = truckmodel.objects.exclude(trans__in=['ISA','THOLADAN','ISA TRANS'])

    if request.method == 'POST' and 'ser1' in request.POST:
        try:
            frdate = request.POST.get('frdate')
            todate = request.POST.get('todate')
            code = request.POST.get('code')

            if (not frdate or frdate == '') and (not todate or todate == '') and (
                    code == '' or code == '----------------------------------Select---------------------------------------'):
                messages.error(request, "Select any filter")
                return render(request, 'staff/out/outtriporderupdate.html', {'code': x})

            filter_kwargs = {}
            if code and code != '----------------------------------Select---------------------------------------':
                filter_kwargs['code'] = code
            if frdate and todate:
                filter_kwargs['dat__range'] = [frdate, todate]

            trips = tripmodel.objects.filter(**filter_kwargs).order_by('dat')

            if trips.exists():
                combined_data = []
                diesel = 0
                hire = 0
                halt = 0

                for trip in trips:
                    trans_data = truckmodel.objects.filter(code=trip.code).exclude(trans__in=['ISA','THOLADAN','ISA TRANS']).first()
                    if trans_data:
                        diesel += trip.dis
                        halt += trip.halt
                        hire += trip.hire
                        combined_data.append({'trip': trip})

                return render(request, 'admin/out/outdailyupdate.html', {'code': x, 'ser1': combined_data, 'diesel': diesel, 'hire': hire, 'halt': halt})
            else:
                messages.error(request, 'Trip not found.')
        except tripmodel.DoesNotExist:
            messages.error(request, 'Trip not found.')

    return render(request, 'admin/out/outdailyupdate.html', {'code': x})

def adminvehicle(request):
    """Renders the admin vehicle page."""
    if 'username' not in request.session:
        return redirect('login')

    today = timezone.localdate()
    previous_trips = None

    try:
        trip = tripmodel.objects.filter(code__trans__in=['ISA','THOLADAN','ISA TRANS']).filter(~Q(dat=today))

        if trip.exists():
            latest_trip = trip.latest('dat')
            latest_day = latest_trip.dat
            diff = (today - latest_day).days

            if diff >= 1:
                day = today - timezone.timedelta(days=diff)
                day = day.strftime('%Y-%m-%d')

                current_trips = tripmodel.objects.filter(dat=today).exclude(place__in=['NO WORK', 'WORKSHOP'])
                previous_trips = tripmodel.objects.filter(dat=day, place__in=['NO WORK', 'WORKSHOP']).exclude(code__in=[t.code for t in current_trips]).order_by('dat')

    except Exception as e:
        print(f"An error occurred: {e}")

    return render(request, 'admin/ISA/adminvehicle.html', {'trucks': previous_trips})

def batta(request, id):
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.filter(trans__in=['ISA', 'THOLADAN', 'ISA TRANS'])
    isa_driver = drivermodel.objects.exclude(driver='OTHER')
    trip = tripmodel.objects.get(pk=id)
    today = datetime.today().strftime('%Y-%m-%d')

    if request.method == 'POST':
        codeid = request.POST.get('code')
        driverid = request.POST.get('driver')
        date = request.POST.get('date')
        battadate = request.POST.get('battadate') or ''
        sheetno = request.POST.get('sheetno') or 0
        opkm = request.POST.get('opkm') or 0
        clkm = request.POST.get('clkm') or 0
        battas = request.POST.get('batta')  or 0
        addbatta = request.POST.get('addbatta') or 0
        genbatta = request.POST.get('genbatta') or 0
        lift = request.POST.get('lift') or 0
        print_val = request.POST.get('print') or 0
        othexp = request.POST.get('othexp') or 0
        wei = request.POST.get('wei') or 0
        halt = request.POST.get('halt') or 0
        park = request.POST.get('park') or 0
        adv = request.POST.get('adv') or 0
        loan = request.POST.get('loan') or 0
        tyrework = request.POST.get('tyrework') or 0
        rto = request.POST.get('rto') or 0
        cont = request.POST.get('cont')
        remarkid = request.POST.get('remarkid') or ''
        remark = request.POST.get('remark') or ''

        try:
            code = truckmodel.objects.get(code=codeid)
            driver = drivermodel.objects.get(driver=driverid)
        except (truckmodel.DoesNotExist, drivermodel.DoesNotExist):
            messages.error(request, "Truck or Driver not found")
            return render(request, 'staff/ISA/batta/batta.html', {'code': isa_trucks, 'driver': isa_driver, 'trip': trip})

        batta_obj = battamodel.objects.create(
            code=code,
            driver=driver,
            date=date,
            battadate=battadate if battadate else None,
            sheetno=sheetno,
            batta=battas,
            addbatta=addbatta,
            lift=lift,
            print=print_val,
            othexp=othexp,
            remarkid=remarkid,
            remark=remark,
            wei=wei,
            halt=halt,
            park=park,
            rto=rto,
            adv=adv,
            tyrework=tyrework,
            genbatta=genbatta,
            loan=loan,
            created_by=request.user
        )

        batta_obj.total = int(battas or 0) + int(genbatta or 0) + int(lift or 0) + int(tyrework or 0) + int(print_val or 0) + int(
            othexp or 0) + int(wei or 0) + int(halt or 0) + int(park or 0) + int(addbatta or 0) + int(rto or 0) - int(adv or 0) - int(
            loan or 0)
        batta_obj.save()

        trip.amount = int(battas or 0) + int(genbatta or 0) + int(lift or 0) + int(print_val or 0) + int(othexp or 0) + int(wei or 0) + int(
            halt or 0) + int(park or 0) + int(rto or 0) + int(adv or 0) + int(loan or 0) + int(tyrework or 0) + int(addbatta or 0)
        trip.batta = batta_obj.batta
        trip.sheetno = sheetno
        trip.opkm = opkm
        trip.clkm = clkm
        trip.driver = driver
        trip.battatotal = batta_obj.total
        trip.battaid = batta_obj.pk
        trip.cont = cont
        trip.updated_by = request.user
        trip.save()

        messages.success(request, "Batta entered successfully!")
        return redirect('triporderupdate')

    return render(request, 'staff/ISA/batta/batta.html', {'code': isa_trucks, 'driver': isa_driver, 'trip': trip, 'today': today})

def battaupdate(request):
    if 'username' not in request.session:
        return redirect('login')
    isa_driver = drivermodel.objects.filter(checks=True).exclude(driver='OTHER')
    isa_trucks = truckmodel.objects.filter(trans__in=['ISA', 'THOLADAN', 'ISA TRANS'])
    index_type = request.GET.get('index_type')
    if index_type == 'admin_index':
        base_template = 'adminindex.html'
    else:
        base_template = 'staffindex.html'
    filter_params = request.session.get('filter_params')
    places = placemodel.objects.all()
    if 'ser1' in request.POST or filter_params:
        try:
            todat = request.POST.get('todate')
            frdat = request.POST.get('frdate')
            if todat or frdat:
                todate = datetime.strptime(todat, '%Y-%m-%d')
                frdate = datetime.strptime(frdat, '%Y-%m-%d')
            else:
                frdate = None
                todate = None
            if 'ser1' in request.POST:
                driver = request.POST.get('driver')
                code = request.POST.get('code')
                place = request.POST.get('place')
                request.session['filter_params'] = {
                    'frdate': frdate.strftime('%Y-%m-%d') if frdate else '',
                    'todate': todate.strftime('%Y-%m-%d') if todate else '',
                    'code': code,
                    'driver': driver,
                    'place': place
                }
            else:
                frdate_str = filter_params.get('frdate')
                todate_str = filter_params.get('todate')
                code = filter_params.get('code')
                driver = filter_params.get('driver')
                place = filter_params.get('place')
                if frdate_str:
                    frdate = datetime.strptime(frdate_str, '%Y-%m-%d')
                else:
                    frdate = None
                if todate_str:
                    todate = datetime.strptime(todate_str, '%Y-%m-%d')
                else:
                    todate = None
            if (not frdate or frdate == '' or frdate == None) and (not todate or todate == '' or todate == None) and ( driver == '--------------------Select-------------------' or '') and ( code == '--------------------Select-------------------' or '') and ( place == '--------------------Select-------------------' or ''):
                messages.error(request, 'Select any filters')
                return render(request, 'staff/ISA/batta/battaupdate.html', {'driver': isa_driver, 'code': isa_trucks,
                                                                    'places':places, 'base_template':base_template})
            filter_kwargs = {}
            if code != '--------------------Select-------------------':
                filter_kwargs['code__code'] = code
            if driver != '--------------------Select-------------------':
                filter_kwargs['driver__driver'] = driver
            if frdate and todate:
                filter_q = Q(battadate__range=[frdate + timedelta(days=2), todate + timedelta(days=2)]) | Q(battadate__isnull=True, date__range=[frdate, todate]) | Q(date__range=[frdate, todate])
            else:
                filter_q = Q()
            if place != '--------------------Select-------------------':
                trip_ids = tripmodel.objects.filter(place=place).values_list('battaid', flat=True)
                batta = battamodel.objects.filter(id__in=trip_ids).filter(filter_q).filter(**{k: v for k, v in filter_kwargs.items() if k not in ['battadate__range', 'date__range']}).order_by('date')
            else:
                batta = battamodel.objects.filter(filter_q).filter(**{k: v for k, v in filter_kwargs.items() if k not in ['battadate__range', 'date__range']}).order_by('date')
            if batta:
                totals = {
                    'batta': str(sum(e.batta for e in batta)),
                    'addbatta': str(sum(e.addbatta for e in batta)),
                    'genbatta': str(sum(e.genbatta for e in batta)),
                    'lift': str(sum(e.lift for e in batta)),
                    'print': str(sum(e.print for e in batta)),
                    'othexp': str(sum(e.othexp for e in batta)),
                    'wei': str(sum(e.wei for e in batta)),
                    'tyrework': str(sum(e.tyrework for e in batta)),
                    'halt': str(sum(e.halt for e in batta)),
                    'park': str(sum(e.park for e in batta)),
                    'rto': str(sum(e.rto for e in batta)),
                    'adv': str(sum(e.adv for e in batta)),
                    'loan': str(sum(e.loan for e in batta)),
                    'total': str(sum(e.total for e in batta)),
                }
                return render(request, 'staff/ISA/batta/battaupdate.html', {'total': totals, 'ser1': batta,
                       'driver': isa_driver, 'code': isa_trucks,'truck':code,'drive':driver, 'frdate': frdate.strftime('%Y-%m-%d')
                        if frdate else '', 'todate': todate.strftime('%Y-%m-%d') if todate else '','places':places, 'place':place,
                                                                            'base_template':base_template})
            else:
                messages.error(request, "Batta not found.")
                return render(request, 'staff/ISA/batta/battaupdate.html', {'ser1': batta, 'driver': isa_driver,
                        'code': isa_trucks,'truck':code,'drive':driver, 'frdate': frdate.strftime('%Y-%m-%d') if frdate else '',
                        'todate': todate.strftime('%Y-%m-%d') if todate else '','places':places, 'place':place, 'base_template':base_template})
        except tripmodel.DoesNotExist:
            messages.error(request, 'No trips found.')
            return render(request, 'staff/ISA/batta/battaupdate.html', {'total': totals, 'ser1': batta, 'driver': isa_driver,
                    'code': isa_trucks,'truck':code,'drive':driver, 'frdate': frdate.strftime('%Y-%m-%d') if frdate else '',
                    'todate': todate.strftime('%Y-%m-%d') if todate else '','places':places, 'place':place, 'base_template':base_template})
    return render(request, 'staff/ISA/batta/battaupdate.html', {'driver': isa_driver,'code': isa_trucks,'places':places,
                                                                 'base_template':base_template})

def battaup(request, id):
    """Renders the batta update page."""
    if 'username' not in request.session:
        return redirect('login')

    isa_driver = drivermodel.objects.filter(checks=True).exclude(driver='OTHER')

    try:
        batta = battamodel.objects.get(pk=id)
        trip = tripmodel.objects.get(battaid=id)
    except (battamodel.DoesNotExist, tripmodel.DoesNotExist):
        messages.error(request, "Batta not found.")
        return redirect('battaupdate')

    if request.method == 'POST':
        batta.sheetno = request.POST.get('sheetno') or 0
        batta.batta = request.POST.get('batta') or 0
        batta.addbatta = request.POST.get('addbatta') or 0
        batta.genbatta=request.POST.get('genbatta') or 0
        batta.lift = request.POST.get('lift') or 0
        batta.driver = drivermodel.objects.get(driver=request.POST.get('driver'))
        trip.opkm = request.POST.get('opkm') or 0
        trip.clkm = request.POST.get('clkm') or 0
        batta.print = request.POST.get('print') or 0
        batta.wei = request.POST.get('wei') or 0
        batta.halt = request.POST.get('halt') or 0
        batta.park = request.POST.get('park') or 0
        batta.rto = request.POST.get('rto') or 0
        batta.adv = request.POST.get('adv') or 0
        batta.tyrework = request.POST.get('tyrework') or 0
        batta.loan = request.POST.get('loan') or 0
        batta.othexp = request.POST.get('othexp') or 0
        batta.remarkid = request.POST.get('remarkid') or ''
        batta.remark = request.POST.get('remark') or ''
        batta.updated_by = request.user
        trip.updated_by = request.user

        # Validate total calculation
        total_fields = [
            batta.batta,
            batta.addbatta,
            batta.genbatta,
            batta.lift,
            batta.print,
            batta.othexp,
            batta.wei,
            batta.halt,
            batta.park,
            batta.rto,
            batta.tyrework
        ]

        try:
            batta.total = (sum(map(int, total_fields))) - int(batta.adv) - int(batta.loan)
            trip.battatotal=batta.total
            trip.batta=batta.batta
            trip.sheetno=batta.sheetno
            trip.amount = (sum(map(int, total_fields))) + int(batta.adv) + int(batta.loan)
            trip.driver = drivermodel.objects.get(driver=request.POST.get('driver'))
            batta.save()
            trip.save()
            messages.success(request, "Batta updated successfully.")
            redirect_url = request.GET.get('redirect')
            if redirect_url == 'triporderupdate':
                return redirect('triporderupdate')
            else:
                return redirect('battaupdate')
        except ValueError:
            messages.error(request, "Invalid input for total calculation.")
            return render(request, 'staff/ISA/batta/battaup.html', {'code': batta, 'trip': trip, 'driver': isa_driver})

    return render(request, 'staff/ISA/batta/battaup.html', {'code': batta, 'trip': trip, 'driver': isa_driver})

def battadel(request, id):
    batta = get_object_or_404(battamodel, pk=id)
    batta.deleted_by=request.user
    batta.delete()
    try:
        trip = tripmodel.objects.get(battaid=id)
        trip.battaid = 0
        trip.sheetno = 0
        trip.opkm = 0
        trip.clkm = 0
        trip.amount = 0
        trip.updated_by=request.user
        trip.save()
    except tripmodel.DoesNotExist:
        pass
    return redirect('battaupdate')

def staffbattaview(request):
    """Renders the staff batta view page."""
    if 'username' not in request.session:
        return redirect('login')

    isa_driver = drivermodel.objects.filter(checks=True).exclude(driver='OTHER')
    isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])

    if request.method == 'POST':
        frdate = request.POST.get('frdate')
        todate = request.POST.get('todate')
        driver = request.POST.get('driver')
        code = request.POST.get('code')

        if (not frdate or frdate == '') and (not todate or todate == '') and (driver == '--------------------Select-------------------') and (code == '--------------------Select-------------------'):
            messages.error(request, 'Select any filters')
            return render(request, 'staff/reports/staffbattaview.html', {'driver': isa_driver, 'code': isa_trucks})

        filter_kwargs = {}
        if code != '--------------------Select-------------------':
            filter_kwargs['code'] = code
        if driver != '--------------------Select-------------------':
            filter_kwargs['driver'] = driver
        if frdate and todate:
            filter_kwargs['date__range'] = [frdate, todate]

        try:
            batta = battamodel.objects.filter(**filter_kwargs).order_by('date')
        except Exception as e:
            messages.error(request, "An error occurred: {}".format(e))
            return render(request, 'staff/reports/staffbattaview.html', {'driver': isa_driver, 'code': isa_trucks})

        if batta:
            totals = {
                'batta': str(sum(e.batta for e in batta)),
                'genbatta': str(sum(e.genbatta for e in batta)),
                'lift': str(sum(e.lift for e in batta)),
                'print': str(sum(e.print for e in batta)),
                'othexp': str(sum(e.othexp for e in batta)),
                'wei': str(sum(e.wei for e in batta)),
                'halt': str(sum(e.halt for e in batta)),
                'park': str(sum(e.park for e in batta)),
                'rto': str(sum(e.rto for e in batta)),
                'adv': str(sum(e.adv for e in batta)),
                'loan': str(sum(e.loan for e in batta)),
                'total': str(sum(e.total for e in batta)),
            }
        else:
            messages.error(request, 'Batta not found')
            totals = {}

        if 'export' in request.POST:
            html = render_to_string('staff/reports/battareport.html', {'total': totals, 'ser1': batta})
            wb = Workbook()
            ws = wb.active
            soup = BeautifulSoup(html, 'html.parser')
            table = soup.find('table')
            rows = table.find_all('tr')

            row_idx = 1
            for row in rows:
                cells = row.find_all('td')
                for i, cell in enumerate(cells):
                    value = cell.text
                    if value.replace('.', '', 1).replace('-', '', 1).isdigit():
                        ws.cell(row=row_idx, column=i + 1).value = float(value)
                        if float(value) < 0:
                            ws.cell(row=row_idx, column=i + 1).value = float(value)
                            ws.cell(row=row_idx, column=i + 1).number_format = '#,##0.00;[Red]-#,##0.00'
                        else:
                            ws.cell(row=row_idx, column=i + 1).value = float(value)
                            ws.cell(row=row_idx, column=i + 1).number_format = '#,##0.00'
                    else:
                        ws.cell(row=row_idx, column=i + 1).value = value
                    if row_idx == 1:
                        ws.cell(row=row_idx, column=i + 1).font = Font(bold=True)
                row_idx += 1

            wb.save('battareport.xlsx')
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'filename="battareport.xlsx"'
            wb.save(response)
            return response

        return render(request, 'staff/reports/staffbattaview.html', {'total': totals, 'ser1': batta, 'driver': isa_driver, 'code': isa_trucks})

    return render(request, 'staff/reports/staffbattaview.html', {'driver': isa_driver, 'code': isa_trucks})

def battahistory(request):
    if 'username' in request.session:
        isa_driver = drivermodel.objects.filter(checks=True).exclude(driver='OTHER')
        isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
        if request.method == 'POST':
            if 'ser1' in request.POST:
                frdate = request.POST.get('frdate')
                todate = request.POST.get('todate')
                driver = request.POST.get('driver')
                code = request.POST.get('code')
                if (not frdate or frdate == '') and (not todate or todate == '') and (
                        driver == '--------------------Select-------------------') and (
                        code == '--------------------Select-------------------'):
                    messages.error(request, 'Select any filters')
                    return render(request, 'admin/history/battahistory.html',
                                    {'driver': isa_driver, 'code': isa_trucks})
                # Rest of your code...
                filter_kwargs = {}

                if code != '--------------------Select-------------------':
                    filter_kwargs['code'] = code

                if driver != '--------------------Select-------------------':
                    filter_kwargs['driver'] = driver

                if frdate and todate:
                    filter_kwargs['date__range'] = [frdate, todate]

                batta = battamodel.objects.filter(**filter_kwargs).order_by('date')
                if batta:
                    totals = {
                        'batta': str(sum(e.batta for e in batta)),
                        'genbatta': str(sum(e.genbatta for e in batta)),
                        'lift': str(sum(e.lift for e in batta)),
                        'print': str(sum(e.print for e in batta)),
                        'othexp': str(sum(e.othexp for e in batta)),
                        'wei': str(sum(e.wei for e in batta)),
                        'halt': str(sum(e.halt for e in batta)),
                        'park': str(sum(e.park for e in batta)),
                        'rto': str(sum(e.rto for e in batta)),
                        'adv': str(sum(e.adv for e in batta)),
                        'loan': str(sum(e.loan for e in batta)),
                        'total': str(sum(e.total for e in batta)),
                    }
                else:
                    messages.error(request, 'Batta not found')
                    totals = {}
                return render(request, 'admin/history/battahistory.html',
                                {'total': totals, 'ser1': batta, 'driver': isa_driver, 'code': isa_trucks})
        return render(request, 'admin/history/battahistory.html',
                        {'driver': isa_driver, 'code': isa_trucks})
    else:
        return redirect('login')

def battaorderhistory(request, id):
    if 'username' in request.session:
        try:
            trip = battamodel.objects.get(pk=id)
            history = trip.history.all()
            history.update(created_at=timezone.localtime(trip.created_at), updated_at=timezone.localtime(trip.updated_at))
            return render(request, 'admin/history/battaorderhistory.html', {'history': history})
        except battamodel.DoesNotExist:
            request.session['error'] = 'Trip not found'
            return render(request, 'admin/history/battaorderhistory.html', {'error': 'Trip not found'})
        except Exception as e:
            request.session['error'] = str(e)
            return render(request, 'admin/history/battaorderhistory.html', {'error': str(e)})
    else:
        return redirect('login')

def delbattahistory(request):
    if 'username' in request.session:
        isa_driver = drivermodel.objects.filter(checks=True).exclude(driver='OTHER')
        isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
        if request.method == 'POST':
            if 'ser1' in request.POST:
                frdate = request.POST.get('frdate')
                todate = request.POST.get('todate')
                driver = request.POST.get('driver')
                code = request.POST.get('code')

                if not (
                        frdate or todate or driver != '--------------------Select-------------------' or code != '--------------------Select-------------------'):
                    messages.error(request, "Select any filter")
                    return render(request, 'admin/history/delbattahistory.html',
                                  {'driver': isa_driver, 'code': isa_trucks})
                filters = {}

                if code != '--------------------Select-------------------':
                    filters['code'] = code

                if driver != '--------------------Select-------------------':
                    filters['driver'] = driver

                if frdate and todate:
                    filters['date__range'] = [frdate, todate]

                batta = battamodel.history.filter(**filters,history_type__in=['-']).order_by('date')
                if batta.exists():
                    return render(request, 'admin/history/delbattahistory.html',
                                {'ser1': batta, 'driver': isa_driver, 'code': isa_trucks})
                else:
                    messages.error(request, "Batta not found")
        return render(request, 'admin/history/delbattahistory.html',
                        {'driver': isa_driver, 'code': isa_trucks})
    else:
        return redirect('login')

def delbattaorderhistory(request, id):
    if 'username' in request.session:
        try:
            history = battamodel.history.filter(id=id)
            return render(request, 'admin/history/delbattaorderhistory.html', {'history': history})
        except battamodel.DoesNotExist:
            request.session['error'] = 'Trip not found'
            return render(request, 'admin/history/delbattaorderhistory.html', {'error': 'Trip not found'})
        except Exception as e:
            request.session['error'] = str(e)
            return render(request, 'admin/history/delbattaorderhistory.html', {'error': str(e)})
    else:
        return redirect('login')

def expense(request, id):
    """Renders the expense page."""
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
    trip = tripmodel.objects.get(pk=id)

    if request.method == 'POST':
        codeid = request.POST.get('code')
        try:
            code = truckmodel.objects.get(code=codeid)
        except truckmodel.DoesNotExist:
            messages.error(request, "Truck not found")
            return redirect('expense')

        date = request.POST.get('date')
        roro = request.POST.get('roro') or 0
        adblue = request.POST.get('adblue') or 0
        oil = request.POST.get('oil') or 0
        insur = request.POST.get('insur') or 0
        tax = request.POST.get('tax') or 0
        taxk = request.POST.get('taxk') or 0
        taxtn = request.POST.get('taxtn') or 0
        welfare = request.POST.get('welfare') or 0
        rtofine = request.POST.get('rtofine') or 0
        test = request.POST.get('test') or 0
        work = request.POST.get('work') or 0
        spare = request.POST.get('spare') or 0
        workshop = request.POST.get('workshop') or ''
        rethread = request.POST.get('rethread') or 0
        tyre = request.POST.get('tyre') or 0
        toll = request.POST.get('toll') or 0
        tyrework = request.POST.get('tyrework') or 0
        othexp = request.POST.get('othexp') or 0
        remark = request.POST.get('remark') or ''
        remarkid = request.POST.get('remarkid') or ''

        try:
            total = int(roro) + int(adblue) + int(oil) + int(insur) + int(tax) + int(test) + int(work) + int(
                spare) + int(rethread) + int(tyre) + int(toll) + int(taxk) + int(taxtn) + int(welfare) + int(rtofine) + int(
                tyrework) + int(othexp)
        except ValueError:
            messages.error(request, "Invalid input for total calculation")
            return render(request, 'staff/ISA/expense/expense.html', {'code': isa_trucks})

        created_by = request.user
        expense = expensemodel.objects.create(
            code=code,
            date=date,
            roro=roro,
            adblue=adblue,
            oil=oil,
            insur=insur,
            tax=tax,
            remark=remark,
            remarkid=remarkid,
            taxk=taxk,
            taxtn=taxtn,
            welfare=welfare,
            rtofine=rtofine,
            test=test,
            work=work,
            spare=spare,
            workshop=workshop,
            rethread=rethread,
            tyre=tyre,
            tyrework=tyrework,
            toll=toll,
            total=total,
            othexp=othexp,
            created_by=created_by
        )
        messages.success(request, "Expense added successfully")
        expense.save()

        if trip:
            trip.exptotal = total
            trip.expid = expense.id
            trip.updated_by = request.user
            trip.save()

        return redirect('triporderupdate')

    return render(request, 'staff/ISA/expense/expense.html', {'code': isa_trucks, 'trip': trip})

def expenseupdate(request):
    """Renders the expense update page."""
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
    index_type = request.GET.get('index_type')
    if index_type == 'admin_index':
        base_template = 'adminindex.html'
    else:
        base_template = 'staffindex.html'

    filter_params = request.session.get('filter_params')

    if 'ser1' in request.POST or filter_params:
        if 'ser1' in request.POST:
            frdate = request.POST.get('frdate')
            todate = request.POST.get('todate')
            code = request.POST.get('code')
            workshop = request.POST.get('workshop')
            request.session['filter_params'] = {
            'frdate': frdate,
            'todate': todate,
            'code': code,
            'workshop': workshop,
            }
        else:
            frdate = filter_params.get('frdate')
            todate = filter_params.get('todate')
            code = filter_params.get('code')
            workshop = filter_params.get('workshop')
        if (not frdate or frdate == '') and (not todate or todate == '') and (
            code == '--------------------Select-------------------') and (not workshop or workshop == ''):
            messages.error(request, 'Select any filter')
            return render(request, 'staff/ISA/expense/expenseupdate.html', {'code': isa_trucks,
                                    'frdate':frdate,'todate':todate,'truck':code,'workshop':workshop,'base_template':base_template})

        filter_kwargs = {}
        if code != '--------------------Select-------------------':
            filter_kwargs['code'] = code

        if workshop:
            filter_kwargs['workshop']=workshop
        if frdate and todate:
            filter_kwargs['date__range'] = [frdate, todate]
        try:
            expenses = expensemodel.objects.filter(**filter_kwargs).order_by('date')
        except Exception as e:
            messages.error(request, "An error occurred: {}".format(e))
            return render(request, 'staff/ISA/expense/expenseupdate.html', {'code': isa_trucks,
                                    'frdate':frdate,'todate':todate,'truck':code,'workshop':workshop,'base_template':base_template})

        if expenses:
            totals = {
                'roro': sum(e.roro for e in expenses),
                'adblue': sum(e.adblue for e in expenses),
                'oil': sum(e.oil for e in expenses),
                'insur': sum(e.insur for e in expenses),
                'tax': sum(e.tax for e in expenses),
                'taxk': sum(e.taxk for e in expenses),
                'taxtn': sum(e.taxtn for e in expenses),
                'rtofine': sum(e.rtofine for e in expenses),
                'welfare': sum(e.welfare for e in expenses),
                'test': sum(e.test for e in expenses),
                'work': sum(e.work for e in expenses),
                'spare': sum(e.spare for e in expenses),
                'tyrework': sum(e.tyrework for e in expenses),
                'rethread': sum(e.rethread for e in expenses),
                'tyre': sum(e.tyre for e in expenses),
                'toll': sum(e.toll for e in expenses),
                'othexp': sum(e.othexp for e in expenses),
                'total': sum(e.total for e in expenses),
            }
        else:
            messages.error(request, "Expense not found.")
            totals = {}

        return render(request, 'staff/ISA/expense/expenseupdate.html', {'code': isa_trucks, 'ser1': expenses,
                                                'total': totals,'frdate':frdate,'todate':todate,'truck':code,'workshop':workshop,
                                                                        'base_template':base_template})
    return render(request, 'staff/ISA/expense/expenseupdate.html', {'code': isa_trucks,'base_template':base_template})

def expup(request, id):
    """Updates an expense."""
    if 'username' not in request.session:
        return redirect('login')

    try:
        expense = expensemodel.objects.get(pk=id)
    except expensemodel.DoesNotExist:
        messages.error(request, "Expense not found.")
        return redirect('expenseupdate')

    trip = tripmodel.objects.get(expid=id)

    if request.method == 'POST':
        expense.roro = request.POST.get('roro') or 0
        expense.adblue = request.POST.get('adblue') or 0
        expense.oil = request.POST.get('oil') or 0
        expense.insur = request.POST.get('insur') or 0
        expense.tax = request.POST.get('tax') or 0
        expense.taxk = request.POST.get('taxk') or 0
        expense.taxtn = request.POST.get('taxtn') or 0
        expense.rtofine = request.POST.get('rtofine') or 0
        expense.welfare = request.POST.get('welfare') or 0
        expense.othexp = request.POST.get('othexp') or 0
        expense.test = request.POST.get('test') or 0
        expense.work = request.POST.get('work') or 0
        expense.workshop = request.POST.get('workshop') or ''
        expense.remarkid = request.POST.get('remarkid') or ''
        expense.remark = request.POST.get('remark') or ''
        expense.spare = request.POST.get('spare') or 0
        expense.rethread = request.POST.get('rethread') or 0
        expense.tyre = request.POST.get('tyre') or 0
        expense.tyrework = request.POST.get('tyrework') or 0
        expense.toll = request.POST.get('toll') or 0
        expense.updated_by = request.user

        # Validate total calculation
        total_fields = [
            expense.roro,
            expense.adblue,
            expense.oil,
            expense.insur,
            expense.tax,
            expense.taxk,
            expense.taxtn,
            expense.test,
            expense.work,
            expense.othexp,
            expense.spare,
            expense.rethread,
            expense.tyre,
            expense.tyrework,
            expense.toll,
            expense.rtofine,
            expense.welfare
        ]
        expense.total = sum(map(int, total_fields))
        expense.save()
        trip.exptotal = expense.total
        trip.updated_by = request.user
        trip.save()
        messages.success(request, "Expense updated successfully.")
        redirect_url = request.GET.get('redirect')
        if redirect_url == 'triporderupdate':
            return redirect('triporderupdate')
        else:
            return redirect('expenseupdate')

    return render(request, 'staff/ISA/expense/expup.html', {'code': expense})

def expdel(request, id):
    """Deletes an expense."""
    expense = get_object_or_404(expensemodel, pk=id)
    expense.deleted_by = request.user
    expense.delete()

    try:
        trip = tripmodel.objects.get(expid=id)
        trip.expid = 0
        trip.exptotal = 0
        trip.updated_by = request.user
        trip.save()
    except tripmodel.DoesNotExist:
        # Log the error or provide a more informative error message
        pass

    messages.success(request, "Expense deleted successfully.")
    return redirect('expenseupdate')

def staffexpview(request):
    """Renders the staff expense view page."""
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])

    if request.method == 'POST':
        frdate = request.POST.get('frdate')
        todate = request.POST.get('todate')
        code = request.POST.get('code')

        if (not frdate or frdate == '') and (not todate or todate == '') and (code == '--------------------Select-------------------'):
            messages.error(request, "Select any filter")
            return render(request, 'staff/reports/staffexpview.html', {'code': isa_trucks})

        filter_kwargs = {}
        if code != '--------------------Select-------------------':
            filter_kwargs['code'] = code
        if frdate and todate:
            filter_kwargs['date__range'] = [frdate, todate]

        try:
            expenses = expensemodel.objects.filter(**filter_kwargs).order_by('date')
        except Exception as e:
            messages.error(request, "An error occurred: {}".format(e))
            return render(request, 'staff/reports/staffexpview.html', {'code': isa_trucks})

        if expenses:
            totals = {
                'roro': sum(e.roro for e in expenses),
                'adblue': sum(e.adblue for e in expenses),
                'oil': sum(e.oil for e in expenses),
                'park': sum(e.park for e in expenses),
                'insur': sum(e.insur for e in expenses),
                'tax': sum(e.tax for e in expenses),
                'emi': sum(e.emi for e in expenses),
                'test': sum(e.test for e in expenses),
                'work': sum(e.work for e in expenses),
                'spare': sum(e.spare for e in expenses),
                'rethread': sum(e.rethread for e in expenses),
                'tyre': sum(e.tyre for e in expenses),
                'toll': sum(e.toll for e in expenses),
                'total': sum(e.total for e in expenses),
            }
        else:
            messages.error(request, "Expense not found.")
            totals = {}

        if 'export' in request.POST:
            html = render_to_string('staff/reports/expreport.html', {'total': totals, 'ser1': expenses})
            wb = Workbook()
            ws = wb.active
            soup = BeautifulSoup(html, 'html.parser')
            table = soup.find('table')
            rows = table.find_all('tr')
            row_idx = 1
            for row in rows:
                cells = row.find_all('td')
                for i, cell in enumerate(cells):
                    value = cell.text
                    if value.replace('.', '', 1).replace('-', '', 1).isdigit():
                        ws.cell(row=row_idx, column=i + 1).value = float(value)
                        if float(value) < 0:
                            ws.cell(row=row_idx, column=i + 1).value = float(value)
                            ws.cell(row=row_idx, column=i + 1).number_format = '#,##0.00;[Red]-#,##0.00'
                        else:
                            ws.cell(row=row_idx, column=i + 1).value = float(value)
                            ws.cell(row=row_idx, column=i + 1).number_format = '#,##0.00'
                    else:
                        ws.cell(row=row_idx, column=i + 1).value = value
                    if row_idx == 1:
                        ws.cell(row=row_idx, column=i + 1).font = Font(bold=True)
                row_idx += 1
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="expensereport.xlsx"'
            wb.save(response)
            return response

        return render(request, 'staff/reports/staffexpview.html', {'ser1': expenses, 'code': isa_trucks, 'total': totals})

    return render(request, 'staff/reports/staffexpview.html', {'code': isa_trucks})

def expensehistory(request):
    if 'username' in request.session:
        isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
        if request.method == 'POST':
            if 'ser1' in request.POST:
                frdate = request.POST.get('frdate')
                todate = request.POST.get('todate')
                code = request.POST.get('code')
                if (not frdate or frdate == '') and (not todate or todate == '') and (code == '--------------------Select-------------------'):
                    messages.error(request, "Select any filter")
                    return render(request, 'admin/history/expensehistory.html', {'code': isa_trucks})
                filter_kwargs = {}

                if code != '--------------------Select-------------------':
                    filter_kwargs['code'] = code

                if frdate and todate:
                    filter_kwargs['date__range'] = [frdate, todate]

                expense = expensemodel.objects.filter(**filter_kwargs).order_by('date')
                if expense:
                    totals = {
                    'roro': sum(e.roro for e in expense),
                    'adblue': sum(e.adblue for e in expense),
                    'oil': sum(e.oil for e in expense),
                    'park': sum(e.park for e in expense),
                    'insur': sum(e.insur for e in expense),
                    'tax': sum(e.tax for e in expense),
                    'emi': sum(e.emi for e in expense),
                    'test': sum(e.test for e in expense),
                    'work': sum(e.work for e in expense),
                    'spare': sum(e.spare for e in expense),
                    'rethread': sum(e.rethread for e in expense),
                    'tyre': sum(e.tyre for e in expense),
                    'toll': sum(e.toll for e in expense),
                    'total': sum(e.total for e in expense),
                    }
                else:
                    messages.error(request, "Expense not found")
                    totals = {}
                return render(request, 'admin/history/expensehistory.html',
                                {'expenses': expense, 'code': isa_trucks, 'total': totals})
        return render(request, 'admin/history/expensehistory.html', {'code': isa_trucks})
    else:
        return redirect('login')

def expenseorderhistory(request, id):
    if 'username' in request.session:
        try:
            trip = expensemodel.objects.get(id=id)
            history = trip.history.all()
            history.update(created_at=timezone.localtime(trip.created_at), updated_at=timezone.localtime(trip.updated_at))
            return render(request, 'admin/history/expenseorderhistory.html', {'history': history})
        except expensemodel.DoesNotExist:
            request.session['error'] = 'Expense not found'
            return render(request, 'admin/history/expenseorderhistory.html', {'error': 'Expense not found'})
        except Exception as e:
            request.session['error'] = str(e)
            return render(request, 'admin/history/expenseorderhistory.html', {'error': str(e)})
    else:
        return redirect('login')

def delexpensehistory(request):
    if 'username' in request.session:
        isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
        if request.method == 'POST':
            if 'ser1' in request.POST:
                frdate = request.POST.get('frdate')
                todate = request.POST.get('todate')
                code = request.POST.get('code')

                if not (frdate or todate or code != '--------------------Select-------------------'):
                    messages.error(request, "Select any filter")
                    return render(request, 'admin/history/delexpensehistory.html', {'code': isa_trucks})

                filters = {}

                if code != '--------------------Select-------------------':
                    filters['code'] = code

                if frdate and todate:
                    filters['date__range'] = [frdate, todate]

                expense = expensemodel.history.filter(**filters,history_type__in=['-']).order_by('date')
                if expense.exists():
                    return render(request, 'admin/history/delexpensehistory.html',
                                {'expenses': expense, 'code': isa_trucks})
                else:
                    messages.error(request, "Expense not found")
        return render(request, 'admin/history/delexpensehistory.html', {'code': isa_trucks})
    else:
        return redirect('login')

def delexpenseorderhistory(request, id):
    if 'username' in request.session:
        try:
            history = expensemodel.history.filter(id=id)
            return render(request, 'admin/history/delexpenseorderhistory.html', {'history': history})
        except expensemodel.DoesNotExist:
            request.session['error'] = 'Expense not found'
            return render(request, 'admin/history/delexpenseorderhistory.html', {'error': 'Expense not found'})
        except Exception as e:
            request.session['error'] = str(e)
            return render(request, 'admin/history/delexpenseorderhistory.html', {'error': str(e)})
    else:
        return redirect('login')

def mileage(request):
    """Renders the mileage page."""
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])

    if request.method == 'POST' and 'ser1' in request.POST:
        try:
            frdate = request.POST.get('frdate')
            todate = request.POST.get('todate')
            code = request.POST.get('code')

            if (not frdate or frdate == '') and (not todate or todate == '') and (code == '' or code == '----------------------------------Select---------------------------------------'):
                messages.error(request, "Select any filter")
                return render(request, 'admin/ISA/mileage.html', {'code': isa_trucks})

            filter_kwargs = {}
            if code and code != '----------------------------------Select---------------------------------------':
                filter_kwargs['code'] = code
            if frdate and todate:
                filter_kwargs['dat__range'] = [frdate, todate]

            trips = tripmodel.objects.filter(**filter_kwargs).order_by('dat')
            if trips.exists():
                combined_data = []
                for trip in trips:
                    trans_data = truckmodel.objects.filter(code=trip.code, trans__in=['ISA','THOLADAN','ISA TRANS']).first()
                    if trans_data:
                        distance = trip.clkm - trip.opkm
                        mileage = int(distance) / trip.disqnt if trip.disqnt != 0 else 0
                        combined_data.append({
                            'trip': trip,
                            'batta': batta,
                            'trans': trans_data,
                            'mileage': mileage,
                            'distance': distance
                        })
                return render(request, 'admin/ISA/mileage.html', {'ser1': combined_data, 'code': isa_trucks})
            else:
                messages.error(request, 'No trips found.')
                return render(request, 'admin/ISA/mileage.html', {'code': isa_trucks})
        except Exception as e:
            messages.error(request, "An error occurred: {}".format(e))
            return render(request, 'admin/ISA/mileage.html', {'code': isa_trucks})

    return render(request, 'admin/ISA/mileage.html', {'code': isa_trucks})

def isaprofit(request):
    """Renders the ISA profit page."""
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
    index_type = request.GET.get('index_type')
    if index_type == 'admin_index':
        base_template = 'adminindex.html'
    else:
        base_template = 'staffindex.html'
    if request.method == 'POST' and 'ser1' in request.POST:
        try:
            frdate = request.POST.get('frdate')
            todate = request.POST.get('todate')
            code = request.POST.get('code')

            if (not frdate or frdate == '') and (not todate or todate == '') and (code == '' or code == '----------------------------------Select---------------------------------------'):
                messages.error(request, "Select any filter")
                return render(request, 'admin/ISA/isaprofit.html', {'code': isa_trucks,'base_template':base_template})

            filter_kwargs = {}
            if code and code != '----------------------------------Select---------------------------------------':
                filter_kwargs['code'] = code
            if frdate and todate:
                filter_kwargs['dat__range'] = [frdate, todate]

            trips = tripmodel.objects.filter(**filter_kwargs).order_by('dat')
            if trips.exists():
                combined_data = []
                total_hire = 0
                total_discount = 0
                total_batta = 0
                total_exp = 0
                total_com = 0

                for trip in trips:
                    trans_data = truckmodel.objects.filter(code=trip.code, trans__in=['ISA','THOLADAN','ISA TRANS']).first()
                    if trans_data:
                        hire = trip.hire + trip.halt
                        profit = hire - trip.dis - trip.amount - trip.exptotal - trip.com
                        combined_data.append({
                            'trip': trip,
                            'bill': hire,
                            'profit': profit
                        })
                        total_hire += hire
                        total_exp += trip.exptotal
                        total_discount += trip.dis
                        total_batta += trip.amount
                        total_com += trip.com

                total_profit = total_hire - total_discount - total_batta - total_exp - total_com
                return render(request, 'admin/ISA/isaprofit.html', {
                    'total_profit': total_profit,
                    'total_hire': total_hire,
                    'ser1': combined_data,
                    'code': isa_trucks,
                    'total_dis': total_discount,
                    'total_batta': total_batta,
                    'total_exp': total_exp,
                    'total_com': total_com,'base_template':base_template
                })
            else:
                messages.error(request, 'No data found.')
                return render(request, 'admin/ISA/isaprofit.html', {'code': isa_trucks,'base_template':base_template})
        except Exception as e:
            messages.error(request, "An error occurred: {}".format(e))
            return render(request, 'admin/ISA/isaprofit.html', {'code': isa_trucks,'base_template':base_template})

    return render(request, 'admin/ISA/isaprofit.html', {'code': isa_trucks,'base_template':base_template})

def isaprofitstaff(request):
    """Renders the ISA profit page."""
    if 'username' not in request.session:
        return redirect('login')

    isa_trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])

    if request.method == 'POST' and 'ser1' in request.POST:
        try:
            frdate = request.POST.get('frdate')
            todate = request.POST.get('todate')
            code = request.POST.get('code')

            if (not frdate or frdate == '') and (not todate or todate == '') and (code == '' or code == '----------------------------------Select---------------------------------------'):
                messages.error(request, "Select any filter")
                return render(request, 'staff/ISA/profit/isaprofitstaff.html', {'code': isa_trucks})

            filter_kwargs = {}
            if code and code != '----------------------------------Select---------------------------------------':
                filter_kwargs['code'] = code
            if frdate and todate:
                filter_kwargs['dat__range'] = [frdate, todate]

            trips = tripmodel.objects.filter(**filter_kwargs).order_by('dat')
            if trips.exists():
                combined_data = []
                total_hire = 0
                total_discount = 0
                total_batta = 0
                total_exp = 0
                total_com = 0

                for trip in trips:
                    trans_data = truckmodel.objects.filter(code=trip.code, trans__in=['ISA','THOLADAN','ISA TRANS']).first()
                    if trans_data:
                        hire = trip.hire + trip.halt
                        profit = hire - trip.dis - trip.amount - trip.exptotal - trip.com
                        combined_data.append({
                            'trip': trip,
                            'bill': hire,
                            'profit': profit
                        })
                        total_hire += hire
                        total_exp += trip.exptotal
                        total_discount += trip.dis
                        total_batta += trip.amount
                        total_com += trip.com

                total_profit = total_hire - total_discount - total_batta - total_exp - total_com
                return render(request, 'staff/ISA/profit/isaprofitstaff.html', {
                    'total_profit': total_profit,
                    'total_hire': total_hire,
                    'ser1': combined_data,
                    'code': isa_trucks,
                    'total_dis': total_discount,
                    'total_batta': total_batta,
                    'total_exp': total_exp,
                    'total_com': total_com
                })
            else:
                messages.error(request, 'No data found.')
                return render(request, 'staff/ISA/profit/isaprofitstaff.html', {'code': isa_trucks})
        except Exception as e:
            messages.error(request, "An error occurred: {}".format(e))
            return render(request, 'staff/ISA/profit/isaprofitstaff.html', {'code': isa_trucks})

    return render(request, 'staff/ISA/profit/isaprofitstaff.html', {'code': isa_trucks})

def fullstaffview(request):
    if 'username' in request.session:
        isa_trucks = truckmodel.objects.all()
        parties = partymodel.objects.all()
        if request.method == 'POST':
            diesel = 0
            hire = 0
            halt = 0
            com=0
            buy=0
            outadv=0
            frdate = request.POST.get('frdate')
            todate = request.POST.get('todate')
            code = request.POST.get('code')
            party = request.POST.get('party')
            if (not frdate or frdate == '') and (not todate or todate == '') and (
                    party == '----------------------------------Select---------------------------------------' or party == '') and (
                    code == '' or code == '----------------------------------Select---------------------------------------'):
                messages.error(request, 'Select any filters')
                return render(request, 'staff/reports/fullstaffview.html', {'code': isa_trucks,'party':parties})
            filter_kwargs = {}

            if code and code != '----------------------------------Select---------------------------------------':
                filter_kwargs['code'] = code

            if party and party != '----------------------------------Select---------------------------------------':
                filter_kwargs['party'] = party

            if frdate and todate:
                filter_kwargs['dat__range'] = [frdate, todate]

            trips = tripmodel.objects.filter(**filter_kwargs).order_by('dat')
            combined = []
            bata = genbatta= lift = print = othexp = wei = halts = park = rto = adv = loan = total = roro = adblue = oil = insur = 0
            tax = emi = test = work = spare = toll = rethread = bal = othexps= taxtn = taxk = welfare = rtofine = tyrework= 0
            if trips.exists():
                for trip in trips:
                    profit = trip.hire + trip.halt - trip.dis - trip.amount - trip.exptotal - trip.com - trip.buy - trip.outadv
                    bal += profit
                    distance = trip.clkm - trip.opkm
                    mileage = int(distance) / trip.disqnt if trip.disqnt != 0 else 0
                    batta = battamodel.objects.filter(pk=trip.battaid)
                    for batta in batta:
                        bata += batta.batta
                        genbatta+=batta.genbatta
                        lift += batta.lift
                        print += batta.print
                        othexp += batta.othexp
                        wei += batta.wei
                        halts += batta.halt
                        park += batta.park
                        rto += batta.rto
                        adv += batta.adv
                        loan += batta.loan
                        total += batta.total
                    exp = expensemodel.objects.filter(pk=trip.expid)
                    for exp in exp:
                        roro += exp.roro
                        adblue += exp.adblue
                        othexps += exp.othexp
                        welfare += exp.welfare
                        rtofine += exp.rtofine
                        tyrework += exp.tyrework
                        oil += exp.oil
                        insur += exp.insur
                        tax += exp.tax
                        taxtn += exp.taxtn
                        taxk += exp.taxk
                        test += exp.test
                        work += exp.work
                        spare += exp.spare
                        toll += exp.toll
                        rethread += exp.rethread
                    diesel += trip.dis
                    halt += trip.halt
                    hire += trip.hire
                    com += trip.com
                    buy += trip.buy
                    outadv += trip.outadv
                    combined.append({'trip': trip, 'batta': batta, 'exp': exp, 'profit': profit, 'mileage': mileage})
            else:
                messages.error(request, 'Report not found')
            if 'ser1' in request.POST:

                return render(request, 'staff/reports/fullstaffview.html', {'trip': combined, 'halt': halt,
                                'diesel': diesel, 'hire': hire,'bata':bata,'genbatta':genbatta,'lift':lift,'print':print,'othexp':othexp,'wei':wei,
                                'halts':halts,'park':park,'rto':rto,'adv':adv,'loan':loan,'total':total,'roro':roro,'code': isa_trucks,
                                'adblue':adblue,'oil':oil,'insur':insur,'tax':tax,'emi':emi,'test':test,'work':work,'spare':spare,
                                'toll':toll,'rethread':rethread,'bal':bal,'com':com,'party':parties,'buy':buy,'outadv':outadv,
                                'othexps':othexps,'taxtn':taxtn,'taxk':taxk,'welfare':welfare,'rtofine':rtofine,'tyrework':tyrework})

            if 'export' in request.POST:
                html = render_to_string('staff/reports/fullreport.html',{'trip': combined, 'diesel': diesel,
                                'hire': hire, 'total': total,'halt': halt,'bata':bata,'genbatta':genbatta,'lift':lift,'print':print,'othexp':othexp,
                                'halts':halts,'park':park,'rto':rto,'adv':adv,'loan':loan,'total':total,'roro':roro,'wei':wei,
                                'adblue':adblue,'oil':oil,'insur':insur,'tax':tax,'emi':emi,'test':test,'work':work,'spare':spare,
                                'toll':toll,'rethread':rethread,'bal':bal,'com':com,'party':parties,'buy':buy,'outadv':outadv,
                                'othexps':othexps,'taxtn':taxtn,'taxk':taxk,'welfare':welfare,'rtofine':rtofine,'tyrework':tyrework})
                wb = Workbook()
                ws = wb.active
                soup = BeautifulSoup(html, 'html.parser')
                table = soup.find('table')
                rows = table.find_all('tr')
                # Download the image

                row_idx = 1  # Initialize row index

                for row in rows:
                    cells = row.find_all('td')
                    for i, cell in enumerate(cells):
                        value = cell.text
                        if value.replace('.', '', 1).replace('-', '', 1).isdigit():  # Check if value is a number
                            ws.cell(row=row_idx, column=i + 1).value = float(value)
                            if float(value) < 0:
                                ws.cell(row=row_idx, column=i + 1).value = float(value)
                                ws.cell(row=row_idx, column=i + 1).number_format = '#,##0.00;[Red]-#,##0.00'
                            else:
                                ws.cell(row=row_idx, column=i + 1).value = float(value)
                                ws.cell(row=row_idx, column=i + 1).number_format = '#,##0.00'
                        else:
                            ws.cell(row=row_idx, column=i + 1).value = value
                        if row_idx == 1:  # Assuming the heading is in the first row
                            ws.cell(row=row_idx, column=i + 1).font = Font(bold=True)
                    row_idx += 1  # Increment row index

                # Save Excel workbook to a file
                wb.save('fullreport.xlsx')

                # Return Excel file as a response
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'filename="fullreport.xlsx"'
                wb.save(response)
                return response
        return render(request, 'staff/reports/fullstaffview.html', {'code': isa_trucks,'party':parties})
    else:
        return redirect('login')

def truckview(request):
    if 'username' not in request.session:
        return redirect('login')

    trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
    outtruck=truckmodel.objects.exclude(trans__in=['ISA','THOLADAN','ISA TRANS'])

    return render(request, 'admin/views/truckview.html', {'trucks': trucks,'outtruck':outtruck})

def truckup(request, id):
    if 'username' not in request.session:
        return redirect('login')

    truck = get_object_or_404(truckmodel, pk=id)
    drivers = drivermodel.objects.all()
    transporters = transportermodel.objects.all()

    if request.method == 'POST':
        truck.trans = get_object_or_404(transportermodel, pk=request.POST.get('trans'))
        truck.driver = get_object_or_404(drivermodel, pk=request.POST.get('driver'))
        truck.feet = request.POST.get('feet')
        truck.save()
        return redirect('truckview')

    return render(request, 'admin/views/truckup.html', {
        'code': truck,
        'driver': drivers,
        'trans': transporters
    })

def stafftruckview(request):
    if 'username' not in request.session:
        return redirect('login')
    if request.session.get('username') == 'Fredy':
        base_template = 'adminindex.html'
    else:
        base_template = 'staffindex.html'

    trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
    outtruck=truckmodel.objects.exclude(trans__in=['ISA','THOLADAN','ISA TRANS'])

    return render(request, 'staff/views/stafftruckview.html', {'trucks': trucks,'outtruck':outtruck,
                                                               'base_template':base_template})

def stafftruckup(request, id):
    if 'username' not in request.session:
        return redirect('login')

    if request.session.get('username') == 'Fredy':
        base_template = 'adminindex.html'
    else:
        base_template = 'staffindex.html'

    truck = get_object_or_404(truckmodel, pk=id)
    drivers = drivermodel.objects.all()
    transporters = transportermodel.objects.all()

    if request.method == 'POST':
        truck.trans = transportermodel.objects.get(pk=request.POST.get('trans'))
        truck.driver = drivermodel.objects.get(pk=request.POST.get('driver'))
        truck.feet = request.POST.get('feet')
        truck.emi = request.POST.get('emi')
        truck.tenure = request.POST.get('tenure') or None
        truck.save()
        return redirect('stafftruckview')

    return render(request, 'staff/views/stafftruckup.html', {
        'code': truck,
        'driver': drivers,
        'trans': transporters,
        'base_template':base_template
    })

def detail(request):
    if 'username' not in request.session:
        return redirect('login')

    if request.session.get('username') == 'Fredy':
        base_template = 'adminindex.html'
    else:
        base_template = 'staffindex.html'
    driver = drivermodel.objects.filter(checks=True)
    party = partymodel.objects.filter(checks=True)
    place = placemodel.objects.filter(checks=True)
    trip = shiftmodel.objects.filter(checks=True)
    trans=transportermodel.objects.filter(checks=True)
    batta=battaregmodel.objects.all()
    return render(request,'staff/views/detail.html',{'driver':driver,'party':party,'place':place,'trip':trip,
                                                     'trans':trans,'batta':batta,'base_template':base_template})

def partyup(request, id):
    if 'username' not in request.session:
        return redirect('login')

    if request.session.get('username') == 'Fredy':
        base_template = 'adminindex.html'
    else:
        base_template = 'staffindex.html'
    try:
        party = partymodel.objects.get(pk=id)
    except partymodel.DoesNotExist:
        return HttpResponseNotFound("Party not found")
    if request.method == 'POST':
        party.party = request.POST.get('party')
        party.add = request.POST.get('add')
        party.short = request.POST.get('short')
        party.save()
        return redirect('detail')
    return render(request, 'staff/views/partyup.html', {'party': party,'base_template':base_template})

def driverup(request,id):
    if 'username' not in request.session:
        return redirect('login')
    try:
        driver = drivermodel.objects.get(pk=id)
    except drivermodel.DoesNotExist:
        return HttpResponseNotFound("Driver not found")
    if request.method == 'POST':
        driver.driver = request.POST.get('driver')
        driver.salary = request.POST.get('salary')
        driver.save()
        return redirect('detail')
    return render(request, 'staff/views/driverup.html', {'driver': driver})

def battaregup(request, id):
    if 'username' not in request.session:
        return redirect('login')
    if request.session.get('username') == 'Fredy':
        base_template = 'adminindex.html'
    else:
        base_template = 'staffindex.html'
    try:
        batta = battaregmodel.objects.get(pk=id)
    except battaregmodel.DoesNotExist:
        return HttpResponseNotFound("Batta not found")
    if request.method == 'POST':
        batta.place = placemodel.objects.get(place=request.POST.get('place'))
        batta.feet = request.POST.get('feet')
        batta.batta = request.POST.get('batta')
        batta.save()
        return redirect('detail')
    return render(request, 'staff/views/battaup.html', {'batta':batta,'base_template':base_template})

def driverdel(request, id):
    try:
        driver = drivermodel.objects.get(pk=id)
        driver.checks = False
        driver.save()

        truck = truckmodel.objects.get(driver=id)
        assign = drivermodel.objects.get(pk='NILL')
        truck.driver = assign
        truck.save()

        return redirect('detail')

    except (drivermodel.DoesNotExist, truckmodel.DoesNotExist):
        messages.error(request, "Driver or truck not found")
        return redirect('detail')

    except Exception as e:
        messages.error(request, "An error occurred: {}".format(e))
        return redirect('detail')

def tripdel(request,id):
    shift = get_object_or_404(shiftmodel, pk=id)
    shift.checks=False
    shift.save()
    return redirect('detail')

def placedel(request,id):
    place = get_object_or_404(placemodel, pk=id)
    place.checks=False
    place.save()
    return redirect('detail')

def partydel(request,id):
    party = get_object_or_404(partymodel, pk=id)
    party.checks=False
    party.save()
    return redirect('detail')

def transdel(request,id):
    trans = get_object_or_404(transportermodel, pk=id)
    trans.checks=False
    trans.save()
    return redirect('detail')

def recycle(request):
    if 'username' not in request.session:
        return redirect('login')
    driver = drivermodel.objects.filter(checks=False)
    party = partymodel.objects.filter(checks=False)
    place = placemodel.objects.filter(checks=False)
    trip = shiftmodel.objects.filter(checks=False)
    trans = transportermodel.objects.filter(checks=False)
    return render(request,'admin/recycle.html',{'driver':driver,'party':party,'place':place,'trip':trip,
                                                'trans':trans})

def partyrestore(request,id):
    if 'username' not in request.session:
        return redirect('login')
    party = get_object_or_404(partymodel, pk=id)
    party.checks = True
    party.save()
    return redirect('recycle')

def placerestore(request,id):
    if 'username' not in request.session:
        return redirect('login')
    place = get_object_or_404(placemodel, pk=id)
    place.checks = True
    place.save()
    return redirect('recycle')

def driverrestore(request,id):
    if 'username' not in request.session:
        return redirect('login')
    driver = get_object_or_404(drivermodel, pk=id)
    driver.checks = True
    driver.save()
    return redirect('recycle')

def shiftrestore(request,id):
    if 'username' not in request.session:
        return redirect('login')
    shift = get_object_or_404(shiftmodel, pk=id)
    shift.checks = True
    shift.save()
    return redirect('recycle')

def transrestore(request,id):
    if 'username' not in request.session:
        return redirect('login')
    trans = get_object_or_404(transportermodel, pk=id)
    trans.checks = True
    trans.save()
    return redirect('recycle')

def reassign(request):
    if 'username' not in request.session:
        return redirect('login')
    truck=truckmodel.objects.all()
    driver=drivermodel.objects.filter(checks=True)
    return render(request,'staff/ISA/trip/reassign.html',{'trucks':truck,'driver':driver})

def assignup(request, id):
    if 'username' not in request.session:
        return redirect('login')
    truck = get_object_or_404(truckmodel, pk=id)
    current_assignment = DriverAssignment.objects.filter(truck=truck, end_date__isnull=True).first()
    available_drivers = drivermodel.objects.filter(checks=True)

    if request.method == 'POST':
        new_driver_id = request.POST.get('driver')
        try:
            new_driver = drivermodel.objects.get(driver=new_driver_id)
        except drivermodel.DoesNotExist:
            return render(request, 'staff/ISA/trip/assignup.html',
                          {'truck': truck, 'driver': available_drivers, 'error': 'Driver not found'})

        # Check if the new driver has an existing truck assignment
        existing_assignment = DriverAssignment.objects.filter(driver=new_driver, end_date__isnull=True).first()
        if existing_assignment:
            existing_assignment.end_date = request.POST.get('date')
            existing_assignment.save()

        # Update the current assignment for the truck
        if current_assignment:
            current_assignment.end_date = request.POST.get('date')
            current_assignment.save()

        # Create a new assignment for the new driver
        new_assignment = DriverAssignment(
            truck=truck,
            driver=new_driver,
            assignment_date=request.POST.get('date')
        )
        new_assignment.save()

        # Update the truck's driver
        truck.driver = new_driver
        truck.save()

        return redirect('reassign')
    return render(request, 'staff/ISA/trip/assignup.html', {'truck': truck, 'driver': available_drivers})

def monthupdate(request):
    if 'username' not in request.session:
        return redirect('login')
    return render(request,'staff/ISA/month/monthupdate.html')

def monthlyprofit(request):
    if 'username' not in request.session:
        return redirect('login')

    current_year = datetime.now().year
    start_year = 2025
    years = range(start_year, current_year + 1)

    trucks = truckmodel.objects.filter(trans__in=['ISA','THOLADAN','ISA TRANS'])
    index_type = request.GET.get('index_type')
    if index_type == 'admin_index':
        base_template = 'adminindex.html'
    else:
        base_template = 'staffindex.html'

    if request.method == 'POST':
        months = request.POST.get('month')
        year = request.POST.get('year')
        code = request.POST.get('code')
        month_words = {
            '1': 'January',
            '2': 'February',
            '3': 'March',
            '4': 'April',
            '5': 'May',
            '6': 'June',
            '7': 'July',
            '8': 'August',
            '9': 'September',
            '10': 'October',
            '11': 'November',
            '12': 'December'
        }

        month = month_words.get(months)
        if not month or not year or not code:
            messages.error(request, 'Please select a month, year, and code.')
            return render(request, 'admin/ISA/monthlyprofit.html', {'years': years, 'truck': trucks,
                                                                    'base_template':base_template})

        try:
            truck=truckmodel.objects.get(code=code)
            salary=truck.driver.salary
            trips = tripmodel.objects.filter(code=code, dat__month=months, dat__year=year)
            hire = battatotal=dis=repair=park=taxk=taxtn=tax=welfare=rtofine=insur=emi=adblue=newtyre=rethread=toll=test=0
            exptotal=0
            open=trips.order_by('dat')
            close=trips.order_by('-dat')
            opkm = 0
            clkm = 0
            for i in open:
                if i.opkm!=0:
                    opkm=i.opkm
                    break
            for i in close:
                if i.clkm!=0:
                    clkm=i.clkm
                    break
            for i in trips:
                hire+=i.hire+i.halt
                dis+=i.dis
                exptotal+=i.exptotal
                batta = battamodel.objects.filter(sheetno=i.sheetno)
                exp=expensemodel.objects.filter(pk=i.expid)

                if batta.exists():
                    for j in batta:
                        battatotal+=j.batta+j.genbatta+j.lift+j.print+j.othexp+j.wei+j.halt+j.park+j.rto+j.adv+j.loan
                        park+=j.park

                if exp.exists():
                    for k in exp:
                        repair+=k.spare
                        toll+=k.toll
                        taxk+=k.taxk
                        taxtn+=k.taxtn
                        tax+=k.tax
                        welfare+=k.welfare
                        rtofine+=k.rtofine
                        emi+=k.emi
                        newtyre+=k.tyre
                        test+=k.test
                        adblue+=k.adblue


            profit=hire-dis-battatotal-salary-exptotal
            return render(request, 'admin/ISA/monthlyprofit.html', {'hire': hire, 'years': years, 'truck': trucks,
                                                'code': code,'month':month,'year':year,'batta':battatotal,'adblue':adblue,'fuel':dis,
                                                'repair':repair,'rethread':rethread,'tyre':newtyre,'park':park,'test':test,'salary':salary,
                                                'toll':toll,'tax':tax,'rto':rtofine,'insur':insur,'loan':emi,'profit':profit,'opkm':opkm,'clkm':clkm,
                                                'base_template':base_template})
        except tripmodel.DoesNotExist:
            messages.error(request, 'No trips found for the selected month, year, and code.')
            return render(request, 'admin/ISA/monthlyprofit.html', {'years': years, 'truck': trucks,'base_template':base_template})

    return render(request, 'admin/ISA/monthlyprofit.html', {'years': years, 'truck': trucks,'base_template':base_template})

def monthprofit(request):
    if 'username' not in request.session:
        return redirect('login')

    current_year = datetime.now().year
    start_year = 2025
    years = range(start_year, current_year + 1)

    if request.session.get('is_superuser'):
        base_template = 'adminindex.html'
    else:
        base_template = 'staffindex.html'

    filter_params = request.session.get('filter_params')
    today = date.today()
    if 'ser1' in request.POST or filter_params:
        if 'ser1' in request.POST:

            months = request.POST.get('month')
            year = request.POST.get('year')
            request.session['filter_params'] = {
                'months': months,
                'year': year,
            }
        else:
            months = filter_params.get('months')
            year = filter_params.get('year')
        month_words = {
            '1': 'January',
            '2': 'February',
            '3': 'March',
            '4': 'April',
            '5': 'May',
            '6': 'June',
            '7': 'July',
            '8': 'August',
            '9': 'September',
            '10': 'October',
            '11': 'November',
            '12': 'December'
        }
        month = month_words.get(months)
        if not month or not year:
            messages.error(request, 'Please select a month, year, and code.')
            return render(request, 'admin/ISA/monthprofit.html', {'years': years,'base_template':base_template})

        trip = (tripmodel.objects.filter(dat__year=year, dat__month=months,
                                         code__trans__in=['ISA', 'THOLADAN', 'ISA TRANS'])
                .values('code__truckno', 'code', 'code__driver__salary', 'code__emi', 'code__tenure')
                .annotate(total_hire=Sum('hire'), total_batta=Sum('amount'),
                          total_dis=Sum('dis'), total_halt=Sum('halt'), total_disqnt=Sum('disqnt'),
                          first_opkm=Min(Case(When(opkm__gt=0, then='opkm'), default=None)),
                          last_clkm=Max('clkm', filter=Q(clkm__gt=0)))
                .annotate(total_emi=Sum('code__emi'))
                .order_by(Case(When(first_opkm__isnull=True, then=1), default=0), 'first_opkm',
                          '-code__driver__salary'))
        exp = expensemodel.objects.filter(date__year=year, date__month=months).values('code').annotate(
            total_rto=Sum('rtofine'),total_adblue=Sum('adblue'),total_repair=Sum('spare'),total_tyre=Sum('tyre'),
            total_rethread=Sum('rethread'),total_toll=Sum('toll'),total_welfare=Sum('welfare'),total_tax=Sum('tax'),
            total_taxk=Sum('taxk'),total_taxtn=Sum('taxtn'),total_insur=Sum('insur'),total_test=Sum('test'),exp_total=Sum('total'),
            total_battery=Sum('work')
        )

        exp_dict = {e['code']: {'total_rto': e['total_rto'], 'total_adblue': e['total_adblue'],'total_repair': e['total_repair'],
                                'total_tyre': e['total_tyre'],'total_rethread': e['total_rethread'],'total_toll': e['total_toll'],
                                'total_welfare': e['total_welfare'],'total_tax': e['total_tax'],'total_taxk': e['total_taxk'],
                                'total_taxtn': e['total_taxtn'],'total_insur': e['total_insur'],'total_test': e['total_test'],
                                'exp_total': e['exp_total'],'total_battery':e['total_battery']
                                } for e in exp}
        batta = battamodel.objects.filter(date__year=year, date__month=months).values('code').annotate(
            total_park=Sum('park'),total_loan=Sum('loan'),total_adv=Sum('adv'),total_total=Sum('total'))
        batta_dict = {
            b['code']: {'total_park': b['total_park'],'total_batta': b['total_loan'] + b['total_adv'] + b['total_total']
            }for b in batta}
        trip_dict = {}
        global_driver_salary_map = {}
        for t in trip:
            print(f"Tenure: {t['code__tenure']}, Today: {today}, EMI: {t['code__emi']}")
            if t['code__tenure'] is None:
                emi = 0
            elif (t['code__tenure'].year, t['code__tenure'].month) >= (int(year),int(months)):
                emi = t['code__emi']
            else:
                emi = 0
            print(f"Calculated EMI: {emi}")
            truck = truckmodel.objects.get(truckno=t['code__truckno'])
            driver_assignments = DriverAssignment.objects.filter(truck=truck)
            current_date = datetime(int(year), int(months), 1).date()

            # Get the driver assignment that is active in the current month
            driver_assignment = driver_assignments.filter(
                Q(end_date__gte=current_date) | Q(end_date__isnull=True),
                assignment_date__lte=current_date
            ).order_by('-assignment_date').first()

            # Get the driver assignments that ended in the previous month
            previous_month = (current_date.replace(day=1) - timedelta(days=1)).replace(day=1)
            previous_driver_assignments = driver_assignments.filter(
                end_date__gte=previous_month,
                end_date__lt=current_date
            )

            # Get the driver assignment that starts in the current month
            next_driver_assignment = driver_assignments.filter(
                assignment_date__gte=current_date,
                assignment_date__lt=datetime(int(year), int(months) + 1, 1).date()
            ).exclude(id=driver_assignment.id if driver_assignment else None).first()

            driver_salaries = {}
            unique_drivers = set()
            driver_salary_total = 0

            assignments_to_check = []

            # Add current and next driver assignments if they belong to this truck
            if driver_assignment and driver_assignment.truck.truckno == t['code__truckno']:
                assignments_to_check.append(("Current Driver", driver_assignment))
            if next_driver_assignment and next_driver_assignment.truck.truckno == t['code__truckno']:
                assignments_to_check.append(("Next Driver", next_driver_assignment))

            # Include previous assignments for this truck
            for assignment in previous_driver_assignments:
                if assignment.truck.truckno == t['code__truckno']:
                    assignments_to_check.append(("Previous Driver", assignment))

            # Avoid duplicates (same driver appearing multiple times for the same truck)
            seen_driver_truck_pairs = set()

            for label, assignment in assignments_to_check:
                pair = (assignment.driver.driver, assignment.truck.truckno)
                if pair in seen_driver_truck_pairs:
                    continue
                seen_driver_truck_pairs.add(pair)

                driver_name = assignment.driver.driver
                driver_salary = assignment.driver.salary
                driver_salaries[f"{driver_name} ({label})"] = driver_salary

                # Add salary only once per unique driver **per truck**
                if driver_name not in unique_drivers:
                    unique_drivers.add(driver_name)
                    driver_salary_total += driver_salary

                # Add salary to global map (for grand total computation)
                if driver_name not in global_driver_salary_map:
                    global_driver_salary_map[driver_name] = driver_salary

            # Prepare display values
            driver_name_display = ', '.join([name.split(' (')[0] for name in driver_salaries.keys()])
            driver_salary_display = ', '.join([f"{name}: {salary}" for name, salary in driver_salaries.items()])

            # Compute truck-level values
            con_value = (t['last_clkm'] or 0) - (t['first_opkm'] or 0)
            mileage_value = con_value / t['total_disqnt'] if t['total_disqnt'] != 0 else 0

            trip_dict[t['code__truckno']] = {
                'code': t['code__truckno'],
                'driver': driver_name_display,
                'total_hire': t['total_hire'] + t['total_halt'],
                'total_batta': batta_dict.get(t['code'], {}).get('total_batta', 0),
                'salary': driver_salary_display,
                'total_salary': driver_salary_total,
                'tenure': t['code__tenure'],
                'total_battery': exp_dict.get(t['code'], {}).get('total_battery', 0),
                'total_repair': exp_dict.get(t['code'], {}).get('total_repair', 0),
                'total_tyre': exp_dict.get(t['code'], {}).get('total_tyre', 0),
                'total_rethread': exp_dict.get(t['code'], {}).get('total_rethread', 0),
                'total_toll': exp_dict.get(t['code'], {}).get('total_toll', 0),
                'total_park': batta_dict.get(t['code'], {}).get('total_park', 0),
                'total_exp': exp_dict.get(t['code'], {}).get('exp_total', 0),
                'total_dis': t['total_dis'],
                'total_disqnt': t['total_disqnt'],
                'total_rto': exp_dict.get(t['code'], {}).get('total_rto', 0),
                'total_welfare': exp_dict.get(t['code'], {}).get('total_welfare', 0),
                'total_tax': exp_dict.get(t['code'], {}).get('total_tax', 0),
                'total_taxk': exp_dict.get(t['code'], {}).get('total_taxk', 0),
                'total_taxtn': exp_dict.get(t['code'], {}).get('total_taxtn', 0),
                'total_insur': exp_dict.get(t['code'], {}).get('total_insur', 0),
                'total_test': exp_dict.get(t['code'], {}).get('total_test', 0),
                'total_emi': emi,
                'total_adblue': exp_dict.get(t['code'], {}).get('total_adblue', 0),
                'first_opkm': t['first_opkm'] or 0,
                'last_clkm': t['last_clkm'] or 0,
                'con': con_value,
                'mileage': mileage_value,
                'total_profit': t['total_hire'] + t['total_halt']
                                - batta_dict.get(t['code'], {}).get('total_batta', 0)
                                - t['total_dis']
                                - exp_dict.get(t['code'], {}).get('exp_total', 0)
                                - emi
                                - driver_salary_total
            }

        # Compute grand totals using **unique drivers globally**
        totals = {
            'hire': sum(value['total_hire'] for value in trip_dict.values()),
            'dis': sum(value['total_dis'] for value in trip_dict.values()),
            'battery': sum(value['total_battery'] for value in trip_dict.values()),
            'tyre': sum(value['total_tyre'] for value in trip_dict.values()),
            'rethread': sum(value['total_rethread'] for value in trip_dict.values()),
            'repair': sum(value['total_repair'] for value in trip_dict.values()),
            'toll': sum(value['total_toll'] for value in trip_dict.values()),
            'park': sum(value['total_park'] for value in trip_dict.values()),
            'exp_total': sum(value['total_exp'] for value in trip_dict.values()),
            'rto': sum(value['total_rto'] for value in trip_dict.values()),
            'welfare': sum(value['total_welfare'] for value in trip_dict.values()),
            'tax': sum(value['total_tax'] for value in trip_dict.values()),
            'taxk': sum(value['total_taxk'] for value in trip_dict.values()),
            'taxtn': sum(value['total_taxtn'] for value in trip_dict.values()),
            'insur': sum(value['total_insur'] for value in trip_dict.values()),
            'test': sum(value['total_test'] for value in trip_dict.values()),
            'emi': sum(value['total_emi'] for value in trip_dict.values()),
            'adblue': sum(value['total_adblue'] for value in trip_dict.values()),
            # Global salary sum (unique driver)
            'salary': sum(global_driver_salary_map.values()),
            # Profit needs adjustment: remove duplicate driver salaries from per-truck sums
            'profit': sum(value['total_profit'] for value in trip_dict.values())
                      + sum(value['total_salary'] for value in trip_dict.values())
                      - sum(global_driver_salary_map.values())
        }
        if 'ser1' in request.POST:
            return render(request, 'admin/ISA/monthprofit.html', {
                'years': years,'base_template': base_template,'data': trip_dict,'today':today,'total': totals,'month':month,'year':year
            })
        if 'export' in request.POST:
            html = render_to_string('admin/ISA/monthprofitreport.html',
                                    {'years': years, 'base_template': base_template, 'data': trip_dict, 'today': today,
                                     'total': totals, 'month': month, 'year': year})
            wb = Workbook()
            ws = wb.active
            soup = BeautifulSoup(html, 'html.parser')
            table = soup.find('table')
            rows = table.find_all('tr')
            row_idx = 1
            for row in rows:
                cells = row.find_all(['th', 'td'])
                for i, cell in enumerate(cells):
                    value = cell.text.strip()
                    ws.cell(row=row_idx, column=i + 1).value = value
                    if row_idx == 1:
                        ws.cell(row=row_idx, column=i + 1).font = Font(bold=True)
                row_idx += 1
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="month_profit_{month}_{year}.xlsx"'
            wb.save(response)
            return response

    return render(request,'admin/ISA/monthprofit.html',{'years': years,'base_template':base_template})

def updatemonth(request):
    if 'username' not in request.session:
        return redirect('login')

    current_year = datetime.now().year
    start_year = 2025
    years = range(start_year, current_year + 1)

    if request.method == 'POST':
        if 'ser1' in request.POST:
            month = request.POST.get('month')
            year = request.POST.get('year')
            return redirect('upmonth', month=month, year=year)
        else:
            messages.error(request, "Invalid request.")

    return render(request, 'staff/ISA/month/updatemonth.html', {'years': years})
